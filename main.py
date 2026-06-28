"""
GEM Tender Logger — Desktop UI  (v4)
--------------------------------------
• Paste one or many GEM listing blocks  → 7 fields parsed instantly
• "Fetch Details" button opens each URL in Chrome via Selenium
  and scrapes ~15 extra fields from the bid document page
• Full editable table view, progress bar, parse log
• Save selected rows to FY-based Excel file

Requirements:
    pip install openpyxl selenium webdriver-manager
    Google Chrome must be installed

Run:
    python gem_tender_ui.py
"""

import os, re, time, threading
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Selenium (imported lazily so app opens even if not installed) ─────────────
_selenium_ok = False
def _try_import_selenium():
    global _selenium_ok
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from webdriver_manager.chrome import ChromeDriverManager
        _selenium_ok = True
        return webdriver, Options, Service, By, WebDriverWait, EC, ChromeDriverManager
    except ImportError:
        return None

# ── Excel ─────────────────────────────────────────────────────────────────────

XL_HEADERS = [
    "S.No", "Bid No", "Bid URL",
    "Ministry", "Department / Buyer", "Organisation", "Office",
    "Category", "Items", "Quantity",
    "Consignee / Location", "Contract Duration",
    "Estimated Value (Rs)", "Evaluation Method",
    "Bid Type", "Bid to RA",
    "EMD Required", "ePBG Required",
    "MII Compliance", "MSE Purchase Preference",
    "MSE Relaxation", "Startup Relaxation",
    "Min Turnover (Lakhs)", "Experience Required (Yrs)",
    "Bid Opening Date", "Start Date", "End Date",
    "Entry Date", "Remarks",
]
COL_WIDTHS = [5,22,40,28,30,24,20,22,30,8,24,14,16,20,14,10,10,10,10,10,10,10,14,14,18,18,18,14,22]

THIN     = Side(style="thin")
BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", start_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DAT_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill("solid", start_color="DCE6F1")


def financial_year(dt):
    return f"{dt.year}-{str(dt.year+1)[-2:]}" if dt.month >= 4 \
           else f"{dt.year-1}-{str(dt.year)[-2:]}"

def xl_path(folder, fy):
    return os.path.join(folder, f"GEM_Tenders_FY_{fy}.xlsx")

def ensure_workbook(path):
    if os.path.exists(path): return
    wb = Workbook(); ws = wb.active; ws.title = "Tenders"
    ws.append(XL_HEADERS)
    for col in range(1, len(XL_HEADERS)+1):
        c = ws.cell(row=1, column=col)
        c.font=HDR_FONT; c.fill=HDR_FILL; c.border=BORDER
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.row_dimensions[1].height=32
    for i,w in enumerate(COL_WIDTHS,1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A2"; wb.save(path)

def xl_append(path, rows):
    wb = load_workbook(path); ws = wb["Tenders"]; snos=[]
    for rec in rows:
        sno = ws.max_row
        ws.append([
            sno,
            rec.get("bid_no",""),      rec.get("bid_url",""),
            rec.get("ministry",""),    rec.get("dept",""),
            rec.get("organisation",""),rec.get("office",""),
            rec.get("category",""),    rec.get("items",""),
            rec.get("quantity",""),    rec.get("location",""),
            rec.get("contract_dur",""),rec.get("est_value",""),
            rec.get("eval_method",""), rec.get("bid_type",""),
            rec.get("bid_to_ra",""),   rec.get("emd",""),
            rec.get("epbg",""),        rec.get("mii",""),
            rec.get("mse_pref",""),    rec.get("mse_relax",""),
            rec.get("startup_relax",""),rec.get("min_turnover",""),
            rec.get("exp_years",""),   rec.get("bid_opening",""),
            rec.get("start_date",""),  rec.get("end_date",""),
            datetime.now().strftime("%d-%m-%Y %H:%M"),
            rec.get("remarks",""),
        ])
        nr=ws.max_row; alt=sno%2==0
        for col in range(1,len(XL_HEADERS)+1):
            c=ws.cell(row=nr,column=col)
            c.font=DAT_FONT; c.border=BORDER
            c.alignment=Alignment(vertical="center",wrap_text=True)
            if alt: c.fill=ALT_FILL
        ws.row_dimensions[nr].height=20; snos.append(sno)
    wb.save(path); return snos


# ── Block parser (listing paste) ──────────────────────────────────────────────

def split_blocks(text):
    parts = re.split(r'(?=BID\s*NO\s*:)', text, flags=re.IGNORECASE)
    return [p.strip() for p in parts if p.strip()]

def parse_one(text):
    r = {}
    m = re.search(r"BID\s*NO\s*:\s*\[([^\]]+)\]\((https?://[^)]+)\)", text, re.I)
    if m:
        r["bid_no"]=m.group(1).strip(); r["bid_url"]=m.group(2).strip()
    else:
        m2 = re.search(r"BID\s*NO\s*:\s*(.+)", text, re.I)
        if m2: r["bid_no"]=m2.group(1).strip()
    m=re.search(r"Items?\s*:\s*(.+)", text, re.I)
    if m: r["items"]=m.group(1).strip()
    m=re.search(r"Quantit[yi]\w*\s*:\s*(\S+)", text, re.I)
    if m: r["quantity"]=m.group(1).strip()
    m=re.search(r"Department\s*(?:Name\s*(?:And\s*Address)?)?\s*:\s*\n(.*?)(?=\n[A-Z]|\Z)", text, re.I|re.DOTALL)
    if m:
        for ln in m.group(1).splitlines():
            ln=ln.strip()
            if ln: r["dept"]=ln; break
    m=re.search(r"Start\s*Date\s*:\s*(.+)", text, re.I)
    if m: r["start_date"]=m.group(1).strip()
    m=re.search(r"End\s*Date\s*:\s*(.+)", text, re.I)
    if m: r["end_date"]=m.group(1).strip()
    return r


# ── Selenium scraper ──────────────────────────────────────────────────────────

def _txt(el):
    try: return el.text.strip()
    except: return ""

def _cell_after(rows, label):
    """Find a <td> whose text matches label and return the next sibling td text."""
    label_l = label.lower()
    for row in rows:
        cells = row.find_elements("tag name", "td")
        for i, c in enumerate(cells):
            if label_l in c.text.lower() and i+1 < len(cells):
                val = cells[i+1].text.strip()
                if val: return val
    return ""

def _find_text(driver, label):
    """Generic: find element whose text contains label, return next sibling or parent text."""
    try:
        # try xpath: element containing label followed by any element
        els = driver.find_elements("xpath",
            f"//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'{label.lower()}')]")
        for el in els:
            # check next sibling
            try:
                sib = el.find_element("xpath", "following-sibling::*[1]")
                t = sib.text.strip()
                if t and t.lower() != label.lower(): return t
            except: pass
            # check parent's text after stripping label
            try:
                parent_text = el.find_element("xpath","..").text
                cleaned = re.sub(re.escape(el.text),"",parent_text,count=1).strip()
                if cleaned and len(cleaned)<200: return cleaned
            except: pass
    except: pass
    return ""

def scrape_bid_page(url, log_fn=None):
    """
    Open the GeM bid document page with Selenium and extract all available fields.
    Returns a dict of extra fields to merge into the record.
    """
    mods = _try_import_selenium()
    if not mods:
        if log_fn: log_fn("err","Selenium not installed. Run: pip install selenium webdriver-manager")
        return {}

    webdriver, Options, Service, By, WebDriverWait, EC, ChromeDriverManager = mods

    opts = Options()
    # Run visible so GeM doesn't block headless browsers
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("--start-maximized")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    # Comment the next line to see the browser window
    # opts.add_argument("--headless=new")

    extra = {}
    driver = None
    try:
        if log_fn: log_fn("info", f"  Opening Chrome → {url}")
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=opts)
        driver.execute_script("Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")

        driver.get(url)
        # wait for page body
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        time.sleep(3)   # let JS render

        page = driver.page_source

        def regex_page(pattern, flags=re.I|re.S):
            m = re.search(pattern, page, flags)
            return m.group(1).strip() if m else ""

        # ── field extractions from page HTML ──────────────────────────────────

        # Bid Opening Date
        extra["bid_opening"] = regex_page(
            r"Bid Opening Date[^<]*?</?\w*[^>]*>\s*([0-9]{2}[-/][0-9]{2}[-/][0-9]{4}[^<]{0,20})")

        # Estimated Bid Value
        extra["est_value"] = regex_page(
            r"Estimated Bid Value[^<]*?</?\w*[^>]*>\s*([0-9][0-9,\.]+)")

        # Evaluation Method
        extra["eval_method"] = regex_page(
            r"Evaluation Method[^<]*?</?\w*[^>]*>\s*([A-Za-z ]+(?:wise|based)[^<]{0,60})")

        # Ministry
        extra["ministry"] = regex_page(
            r"Ministry[/\w\s]*?Name[^<]*?</?\w*[^>]*>\s*([A-Z][^<]{3,80})")

        # Department
        dept_scraped = regex_page(
            r"Department Name[^<]*?</?\w*[^>]*>\s*([A-Z][^<]{3,80})")
        if dept_scraped: extra["dept"] = dept_scraped

        # Organisation
        extra["organisation"] = regex_page(
            r"Organisation Name[^<]*?</?\w*[^>]*>\s*([A-Z][^<]{2,60})")

        # Office
        extra["office"] = regex_page(
            r"Office Name[^<]*?</?\w*[^>]*>\s*([A-Z][^<]{2,60})")

        # Contract Duration
        extra["contract_dur"] = regex_page(
            r"Contract (?:Duration|Period)[^<]*?</?\w*[^>]*>\s*([0-9][^<]{1,30})")

        # Bid Type
        extra["bid_type"] = regex_page(
            r"Type of Bid[^<]*?</?\w*[^>]*>\s*([A-Za-z ]+Bid[^<]{0,40})")

        # Bid to RA
        ra = regex_page(r"Bid to RA enabled[^<]*?</?\w*[^>]*>\s*(Yes|No)")
        extra["bid_to_ra"] = ra if ra else regex_page(r"Bid to RA\s*</?\w*[^>]*>\s*(Yes|No)")

        # EMD
        emd = regex_page(r"EMD Detail[^<]{0,200}?Required[^<]*?</?\w*[^>]*>\s*(Yes|No)")
        extra["emd"] = emd

        # ePBG
        epbg = regex_page(r"ePBG Detail[^<]{0,200}?Required[^<]*?</?\w*[^>]*>\s*(Yes|No)")
        extra["epbg"] = epbg

        # MII Compliance
        extra["mii"] = regex_page(
            r"MII Compliance[^<]*?</?\w*[^>]*>\s*(Yes|No)")

        # MSE Purchase Preference
        extra["mse_pref"] = regex_page(
            r"MSE Purchase Preference[^<]*?</?\w*[^>]*>\s*(Yes|No)")

        # MSE Relaxation
        extra["mse_relax"] = regex_page(
            r"MSE Relaxation[^<]{0,60}?</?\w*[^>]*>\s*(Yes|No)")

        # Startup Relaxation
        extra["startup_relax"] = regex_page(
            r"Startup Relaxation[^<]{0,60}?</?\w*[^>]*>\s*(Yes|No)")

        # Min Turnover
        extra["min_turnover"] = regex_page(
            r"(?:Minimum Average Annual Turnover|Average Turn Over)[^<]{0,60}?</?\w*[^>]*>\s*([0-9][^<]{0,20})")

        # Experience years
        extra["exp_years"] = regex_page(
            r"Years of Past Experience[^<]{0,60}?</?\w*[^>]*>\s*([0-9][^<]{0,20})")

        # Consignee / delivery location
        loc = regex_page(
            r"Consignee[^<]{0,100}?Address[^<]*?</?\w*[^>]*>\s*([A-Z][^<]{5,120})")
        if not loc:
            loc = regex_page(r"Delivery Location[^<]*?</?\w*[^>]*>\s*([A-Z][^<]{5,80})")
        extra["location"] = loc

        # Category (item category)
        cat = regex_page(r"Item Category[^<]*?</?\w*[^>]*>\s*([A-Z][^<]{3,80})")
        if not cat:
            cat = regex_page(r"Similar Category[^<]*?</?\w*[^>]*>\s*([A-Z][^<]{3,80})")
        extra["category"] = cat

        # Remove empty values
        extra = {k:v for k,v in extra.items() if v}

        if log_fn:
            log_fn("ok", f"  Scraped {len(extra)} extra fields: {', '.join(extra.keys())}")

    except Exception as e:
        if log_fn: log_fn("err", f"  Selenium error: {e}")
    finally:
        if driver:
            try: driver.quit()
            except: pass

    return extra


# ── Theme ─────────────────────────────────────────────────────────────────────

BG      = "#0D1117"
PANEL   = "#161B22"
CARD    = "#21262D"
ACCENT  = "#238636"
ACCENT2 = "#1F6FEB"
MUTED   = "#8B949E"
TEXT    = "#E6EDF3"
TEXTSUB = "#7D8590"
SUCCESS = "#3FB950"
ERR     = "#F85149"
WARN    = "#D29922"
SEL_BG  = "#1F6FEB"

FH = ("Consolas", 9)
FB = ("Segoe UI", 10)
FL = ("Segoe UI", 9)
FT = ("Segoe UI", 12, "bold")


# ── Table columns ─────────────────────────────────────────────────────────────

TV_COLS = [
    ("bid_no",       "Bid No",            180),
    ("ministry",     "Ministry",          160),
    ("dept",         "Department",        180),
    ("organisation", "Organisation",      140),
    ("category",     "Category",          120),
    ("items",        "Items",             160),
    ("quantity",     "Qty",                55),
    ("location",     "Consignee/Location",150),
    ("est_value",    "Est. Value (Rs)",   110),
    ("eval_method",  "Eval Method",       130),
    ("bid_type",     "Bid Type",          110),
    ("bid_to_ra",    "RA",                 45),
    ("emd",          "EMD",                45),
    ("epbg",         "ePBG",               45),
    ("mii",          "MII",                40),
    ("mse_pref",     "MSE Pref",           60),
    ("contract_dur", "Contract Dur",       90),
    ("min_turnover", "Turnover (L)",       80),
    ("exp_years",    "Exp Yrs",            55),
    ("bid_opening",  "Bid Opening",       130),
    ("start_date",   "Start Date",        130),
    ("end_date",     "End Date",          130),
    ("remarks",      "Remarks",           120),
]
TV_IDS = [c[0] for c in TV_COLS]


# ── App ───────────────────────────────────────────────────────────────────────

class TenderApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("GEM Tender Logger  v4")
        self.configure(bg=BG)
        self.geometry("1400x860")
        self.minsize(900, 600)

        self.save_folder = tk.StringVar(
            value=os.path.expanduser("~\\Documents" if os.name=="nt" else "~/Documents"))
        self._records  = []
        self._editing  = None
        self._fetch_running = False

        self._style()
        self._build()
        self._fy_tick()

    # ── styles ────────────────────────────────────────────────────────────────
    def _style(self):
        s = ttk.Style(self); s.theme_use("clam")
        s.configure(".", background=BG, foreground=TEXT, font=FB)
        s.configure("Treeview", background=PANEL, foreground=TEXT,
                    fieldbackground=PANEL, rowheight=26, borderwidth=0, font=FL)
        s.configure("Treeview.Heading", background=CARD, foreground=TEXT,
                    font=("Segoe UI",9,"bold"), relief="flat")
        s.map("Treeview",
              background=[("selected",SEL_BG)], foreground=[("selected",TEXT)])
        s.map("Treeview.Heading", background=[("active",CARD)])
        for o in ("Vertical","Horizontal"):
            s.configure(f"{o}.TScrollbar", background=CARD, troughcolor=BG,
                        bordercolor=BG, arrowcolor=MUTED, gripcount=0)
        s.configure("TProgressbar", troughcolor=CARD, background=ACCENT2,
                    bordercolor=BG, lightcolor=ACCENT2, darkcolor=ACCENT2)

    # ── layout ────────────────────────────────────────────────────────────────
    def _build(self):
        # top bar
        top = tk.Frame(self, bg=PANEL, pady=8, padx=14,
                       highlightthickness=1, highlightbackground="#30363D")
        top.pack(fill="x")
        tk.Label(top, text="GEM Tender Logger", font=FT, bg=PANEL, fg=TEXT).pack(side="left")
        self.fy_lbl = tk.Label(top, text="", font=FL, bg=ACCENT2, fg=TEXT, padx=10, pady=3)
        self.fy_lbl.pack(side="right", padx=(6,0))

        # folder bar
        fbar = tk.Frame(self, bg=BG, pady=5, padx=14)
        fbar.pack(fill="x")
        tk.Label(fbar, text="Save to:", font=FL, bg=BG, fg=MUTED).pack(side="left")
        tk.Entry(fbar, textvariable=self.save_folder, bg=CARD, fg=TEXT,
                 insertbackground=TEXT, relief="flat", font=FL,
                 highlightthickness=1, highlightbackground="#30363D",
                 highlightcolor=ACCENT2, width=60).pack(side="left", padx=(6,4))
        self._btn(fbar, "Browse", self._browse, bg=CARD).pack(side="left")

        # paned: left = paste+log, right = table
        pane = tk.PanedWindow(self, orient="horizontal", bg=BG,
                              sashwidth=5, sashrelief="flat", handlesize=0)
        pane.pack(fill="both", expand=True, padx=10, pady=(4,0))

        # ── LEFT ─────────────────────────────────────────────────────────────
        left = tk.Frame(pane, bg=BG)
        pane.add(left, minsize=320, width=400)

        tk.Label(left, text="Paste GEM Tender Block(s)", font=("Segoe UI",9,"bold"),
                 bg=BG, fg=MUTED).pack(anchor="w", pady=(4,2))
        self.paste_txt = tk.Text(left, bg=CARD, fg=TEXT, insertbackground=TEXT,
                                 relief="flat", font=FH, height=12,
                                 highlightthickness=1, highlightbackground="#30363D",
                                 highlightcolor=ACCENT2, wrap="word", undo=True)
        self.paste_txt.pack(fill="x")
        tk.Label(left, text="Paste one or many blocks — each starting with BID NO:",
                 font=("Segoe UI",8), bg=BG, fg=TEXTSUB).pack(anchor="w", pady=(2,4))

        # progress
        self.progress = ttk.Progressbar(left, orient="horizontal",
                                        mode="determinate", length=200)
        self.progress.pack(fill="x", pady=(0,4))
        self.prog_lbl = tk.Label(left, text="", font=("Segoe UI",8),
                                 bg=BG, fg=TEXTSUB)
        self.prog_lbl.pack(anchor="w")

        # buttons
        btn_row = tk.Frame(left, bg=BG)
        btn_row.pack(fill="x", pady=(4,8))
        self._btn(btn_row, "  1. Parse Blocks  ", self._do_parse,
                  bg=ACCENT2, pad=10).pack(side="left")
        self._btn(btn_row, "  2. Fetch Details (Selenium)  ",
                  self._do_fetch_all, bg=ACCENT, pad=10).pack(side="left", padx=6)
        self._btn(btn_row, "Clear", lambda: self.paste_txt.delete("1.0","end"),
                  bg=CARD).pack(side="left")

        # log
        tk.Label(left, text="Log", font=("Segoe UI",9,"bold"),
                 bg=BG, fg=MUTED).pack(anchor="w", pady=(0,2))
        log_fr = tk.Frame(left, bg=CARD,
                          highlightthickness=1, highlightbackground="#30363D")
        log_fr.pack(fill="both", expand=True)
        self.log_txt = tk.Text(log_fr, bg=CARD, fg=TEXTSUB, relief="flat",
                               font=("Consolas",8), state="disabled", wrap="word",
                               highlightthickness=0)
        lsb = ttk.Scrollbar(log_fr, orient="vertical", command=self.log_txt.yview)
        self.log_txt.configure(yscrollcommand=lsb.set)
        lsb.pack(side="right", fill="y")
        self.log_txt.pack(fill="both", expand=True, padx=4, pady=4)
        for tag,col in [("ok",SUCCESS),("warn",WARN),("err",ERR),("info",ACCENT2)]:
            self.log_txt.tag_configure(tag, foreground=col)

        # ── RIGHT ────────────────────────────────────────────────────────────
        right = tk.Frame(pane, bg=BG)
        pane.add(right, minsize=500)

        hdr = tk.Frame(right, bg=BG)
        hdr.pack(fill="x", pady=(4,4))
        tk.Label(hdr, text="Parsed Tenders", font=("Segoe UI",9,"bold"),
                 bg=BG, fg=MUTED).pack(side="left")
        self.count_lbl = tk.Label(hdr, text="0 rows", font=FL, bg=BG, fg=TEXTSUB)
        self.count_lbl.pack(side="left", padx=8)

        self._btn(hdr,"Select All",   self._sel_all,    bg=CARD).pack(side="right",padx=2)
        self._btn(hdr,"Delete Selected", self._del_sel, bg=CARD, fg=ERR).pack(side="right",padx=2)
        self._btn(hdr,"  Fetch Selected (Selenium)  ",
                  self._do_fetch_sel, bg="#2D333B").pack(side="right",padx=2)
        self._btn(hdr,"  Save Selected to Excel  ",
                  self._save_selected, bg=ACCENT, pad=10).pack(side="right",padx=(6,2))

        # treeview
        tv_fr = tk.Frame(right, bg=BG)
        tv_fr.pack(fill="both", expand=True)
        cols = [c[0] for c in TV_COLS]
        self.tv = ttk.Treeview(tv_fr, columns=cols, show="headings",
                               selectmode="extended")
        for cid, lbl, w in TV_COLS:
            self.tv.heading(cid, text=lbl, command=lambda c=cid: self._sort(c))
            self.tv.column(cid, width=w, minwidth=40, stretch=False)

        vsb = ttk.Scrollbar(tv_fr, orient="vertical",   command=self.tv.yview)
        hsb = ttk.Scrollbar(tv_fr, orient="horizontal", command=self.tv.xview)
        self.tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tv.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tv_fr.rowconfigure(0, weight=1); tv_fr.columnconfigure(0, weight=1)

        self.tv.tag_configure("alt",     background="#1C2128")
        self.tv.tag_configure("fetched", background="#1A2E1A", foreground=SUCCESS)
        self.tv.tag_configure("saved",   background="#1A3A2A", foreground=SUCCESS)
        self.tv.tag_configure("fetching",background="#2A2A1A", foreground=WARN)

        self.tv.bind("<Double-1>", self._on_dbl)
        self.tv.bind("<Button-1>", self._cancel_edit)

        # status bar
        self.status = tk.Label(self, text="Ready", font=FL,
                               bg=PANEL, fg=MUTED, anchor="w", padx=10, pady=4)
        self.status.pack(fill="x", side="bottom")

    # ── widget helpers ────────────────────────────────────────────────────────
    def _btn(self, parent, text, cmd, bg=CARD, fg=TEXT, pad=6):
        return tk.Button(parent, text=text, command=cmd,
                         bg=bg, fg=fg, relief="flat", font=FL,
                         padx=pad, pady=4,
                         activebackground=ACCENT2, activeforeground=TEXT,
                         cursor="hand2")

    def _log(self, level, msg):
        self.log_txt.configure(state="normal")
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_txt.insert("end", f"[{ts}] {msg}\n", level)
        self.log_txt.see("end")
        self.log_txt.configure(state="disabled")
        self.update_idletasks()

    def _set_status(self, msg, color=MUTED):
        self.status.configure(text=msg, fg=color)
        self.update_idletasks()

    def _set_prog(self, val, label=""):
        self.progress["value"] = val
        self.prog_lbl.configure(text=label)
        self.update_idletasks()

    def _fy_tick(self):
        self.fy_lbl.configure(text=f"FY {financial_year(datetime.now())}")
        self.after(60000, self._fy_tick)

    def _browse(self):
        d = filedialog.askdirectory(title="Select output folder",
                                    initialdir=self.save_folder.get())
        if d: self.save_folder.set(d)

    # ── Step 1 — parse paste blocks ───────────────────────────────────────────
    def _do_parse(self):
        raw = self.paste_txt.get("1.0","end").strip()
        if not raw: self._log("warn","Paste area is empty."); return
        self._log("info", f"--- Parse started {datetime.now().strftime('%H:%M:%S')} ---")
        self._set_prog(0,"Splitting blocks…")

        def worker():
            blocks = split_blocks(raw)
            total  = len(blocks)
            self._log("info", f"Found {total} block(s).")
            recs = []
            for i, blk in enumerate(blocks,1):
                rec = parse_one(blk)
                self._set_prog(int(i/total*100), f"Parsing {i}/{total}…")
                time.sleep(0.04)
                if rec.get("bid_no"):
                    recs.append(rec)
                    flds = [k for k,v in rec.items() if v]
                    self._log("ok", f"[{i}/{total}] {rec['bid_no']} — {len(flds)} fields")
                else:
                    self._log("warn", f"[{i}/{total}] SKIP — BID NO not found")
            self.after(0, lambda: self._add_rows(recs, total))
        threading.Thread(target=worker, daemon=True).start()

    def _add_rows(self, recs, total):
        for rec in recs:
            self._records.append(rec)
            self._tv_insert(rec)
        self._refresh_alt()
        self.count_lbl.configure(text=f"{len(self._records)} rows")
        skipped = total - len(recs)
        msg = f"Parsed {total} block(s): {len(recs)} added" + (f", {skipped} skipped" if skipped else "")
        self._log("info", msg)
        self._set_status(msg, SUCCESS if recs else WARN)
        self._set_prog(100, "Done.")
        if recs: self.paste_txt.delete("1.0","end")

    # ── Step 2 — Selenium fetch ───────────────────────────────────────────────
    def _do_fetch_all(self):
        """Fetch details for ALL records that have a bid_url."""
        targets = [(i, r) for i,r in enumerate(self._records) if r.get("bid_url")]
        self._run_fetch(targets, "all")

    def _do_fetch_sel(self):
        """Fetch details for SELECTED rows only."""
        sel = self.tv.selection()
        if not sel:
            self._log("warn","No rows selected."); return
        targets = []
        for iid in sel:
            idx = self.tv.index(iid)
            if idx < len(self._records) and self._records[idx].get("bid_url"):
                targets.append((idx, self._records[idx]))
        if not targets:
            self._log("warn","Selected rows have no Bid URL."); return
        self._run_fetch(targets, "selected")

    def _run_fetch(self, targets, label):
        if self._fetch_running:
            self._log("warn","A fetch is already running. Please wait."); return
        if not targets:
            self._log("warn","No rows with Bid URL to fetch."); return

        if not _try_import_selenium():
            messagebox.showerror("Missing library",
                "Please install Selenium:\n\npip install selenium webdriver-manager")
            return

        self._fetch_running = True
        self._log("info", f"--- Selenium fetch started: {len(targets)} {label} row(s) ---")
        self._set_prog(0, f"Fetching 0/{len(targets)}…")

        iid_map = {i: iid for i,iid in enumerate(self.tv.get_children())}

        def worker():
            total = len(targets)
            for n, (idx, rec) in enumerate(targets, 1):
                iid = iid_map.get(idx)
                bid = rec.get("bid_no","?")
                url = rec["bid_url"]

                self._set_prog(int((n-1)/total*100), f"Fetching {n}/{total}: {bid}")
                self._log("info", f"[{n}/{total}] Fetching {bid}")

                # mark row as "fetching"
                if iid: self.after(0, lambda i=iid: self.tv.item(i, tags=("fetching",)))

                extra = scrape_bid_page(url, log_fn=self._log)

                if extra:
                    rec.update(extra)
                    # update treeview cells
                    def update_tv(i=iid, r=rec):
                        if i:
                            for cid in TV_IDS:
                                if cid in r:
                                    self.tv.set(i, cid, r[cid])
                            self.tv.item(i, tags=("fetched",))
                    self.after(0, update_tv)
                    self._log("ok", f"[{n}/{total}] {bid} — merged {len(extra)} extra fields")
                else:
                    if iid: self.after(0, lambda i=iid: self.tv.item(i, tags=()))
                    self._log("warn", f"[{n}/{total}] {bid} — no extra data scraped")

                self._set_prog(int(n/total*100), f"Fetched {n}/{total}")

            self._fetch_running = False
            self._set_prog(100, "Fetch complete.")
            msg = f"Selenium fetch done: {total} URL(s) processed"
            self._log("info", f"--- {msg} ---")
            self.after(0, lambda: self._set_status(msg, SUCCESS))

        threading.Thread(target=worker, daemon=True).start()

    # ── table helpers ─────────────────────────────────────────────────────────
    def _tv_insert(self, rec):
        vals = tuple(rec.get(c,"") for c in TV_IDS)
        return self.tv.insert("","end", values=vals)

    def _refresh_alt(self):
        for i, iid in enumerate(self.tv.get_children()):
            cur_tags = self.tv.item(iid,"tags")
            special = [t for t in cur_tags if t in ("fetched","saved","fetching")]
            if not special:
                self.tv.item(iid, tags=("alt",) if i%2 else ())

    _sort_state = {}
    def _sort(self, col):
        rev = self._sort_state.get(col, False)
        items = [(self.tv.set(k,col),k) for k in self.tv.get_children()]
        items.sort(reverse=rev)
        for i,(_,k) in enumerate(items): self.tv.move(k,"",i)
        self._sort_state[col] = not rev
        self._refresh_alt()

    def _on_dbl(self, event):
        region = self.tv.identify_region(event.x, event.y)
        if region != "cell": return
        col = self.tv.identify_column(event.x)
        iid = self.tv.identify_row(event.y)
        if not iid: return
        col_idx = int(col[1:])-1
        if col_idx >= len(TV_IDS): return
        col_id = TV_IDS[col_idx]
        bbox = self.tv.bbox(iid, col)
        if not bbox: return
        x,y,w,h = bbox
        var = tk.StringVar(value=self.tv.set(iid, col_id))
        e = tk.Entry(self.tv, textvariable=var, bg=SEL_BG, fg=TEXT,
                     insertbackground=TEXT, relief="flat", font=FL,
                     highlightthickness=0)
        e.place(x=x,y=y,width=w,height=h); e.focus_set(); e.select_range(0,"end")
        self._editing = (iid, col_id, e)
        def commit(ev=None):
            nv = var.get(); self.tv.set(iid,col_id,nv)
            idx = self.tv.index(iid)
            if idx < len(self._records): self._records[idx][col_id]=nv
            e.destroy(); self._editing=None
        e.bind("<Return>",commit); e.bind("<Tab>",commit)
        e.bind("<Escape>",lambda ev: e.destroy())

    def _cancel_edit(self, event):
        if self._editing:
            try: self._editing[2].destroy()
            except: pass
            self._editing=None

    def _sel_all(self):
        self.tv.selection_set(self.tv.get_children())

    def _del_sel(self):
        sel = self.tv.selection()
        if not sel: return
        for iid in sorted(sel, key=self.tv.index, reverse=True):
            idx = self.tv.index(iid)
            self.tv.delete(iid)
            if idx < len(self._records): self._records.pop(idx)
        self._refresh_alt()
        self.count_lbl.configure(text=f"{len(self._records)} rows")
        self._log("info", f"Deleted {len(sel)} row(s).")

    def _save_selected(self):
        sel = self.tv.selection()
        if not sel:
            self._log("warn","No rows selected."); return
        folder = self.save_folder.get().strip()
        if not folder:
            self._set_status("Set output folder first.", ERR); return
        os.makedirs(folder, exist_ok=True)
        fy   = financial_year(datetime.now())
        path = xl_path(folder, fy)
        ensure_workbook(path)
        recs = []
        for iid in sel:
            idx = self.tv.index(iid)
            if idx < len(self._records): recs.append(self._records[idx])
        try:
            snos = xl_append(path, recs)
            msg = f"Saved {len(snos)} row(s) → {os.path.basename(path)}  (S.No {snos[0]}–{snos[-1]})"
            self._log("ok", msg); self._set_status(msg, SUCCESS); self._fy_tick()
            for iid in sel: self.tv.item(iid, tags=("saved",))
        except Exception as ex:
            self._log("err", f"Save error: {ex}")
            messagebox.showerror("Save Error", str(ex))


# ── main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = TenderApp()
    app.mainloop()