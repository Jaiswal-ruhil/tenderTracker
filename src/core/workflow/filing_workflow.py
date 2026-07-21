"""
filing_workflow.py
~~~~~~~~~~~~~~~~~~
Enhanced workflow for tender filing process with automation upgrades:
1. Download tender PDF (if not present) with retry logic
2. Extract required documents from PDF using hybrid regex/LLM approach
3. Match with pre-uploaded firm documents using AI-powered matching
4. Copy documents to filing folder with validation
5. Generate checklist and summary with GEM requirements mapping
6. Automated document validation and integrity checks
"""

import os
import re
import shutil
import json
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import logger
import db
import scraper
import pdf_extractor
from parser import convert_pdf_text_to_markdown
import llm
from excel import financial_year
from alert_system import alert_document_issue, alert_filing_issue, alert_network_issue, AlertSeverity
from document_matcher import DocumentMatcherMixin
import template_generator


class FilingWorkflow(DocumentMatcherMixin):
    """Handles the complete tender filing workflow."""
    
    def __init__(self, log_fn=None, progress_cb=None):
        """
        Initialize the filing workflow with enhanced automation.
        
        Args:
            log_fn: Optional logging function (signature: log_fn(level, message))
            progress_cb: Optional progress callback function (signature: progress_cb(percent, message))
        """
        self.log_fn = log_fn or logger.log
        self.progress_cb = progress_cb
        self.required_documents = []
        self.matched_documents = {}
        self.missing_documents = []
        self.filing_folder = ""
        self._local_llm_used = False
        self.category = "General"
        self.emd_details = {}
        self.validation_results = {}  # Store document validation results
        self.processing_stats = {}  # Track processing statistics
        self.gem_requirements = []
    
    def _validate_file_path(self, path: str) -> bool:
        """Validate if a file path is safe and accessible."""
        if not path or not isinstance(path, str):
            return False
        try:
            # Check for invalid characters
            invalid_chars = '<>:"|?*'
            if any(char in path for char in invalid_chars):
                return False
            # Check if path is too long
            if len(path) > 260:  # Windows MAX_PATH
                return False
            return True
        except Exception:
            return False
    
    def _validate_url(self, url: str) -> bool:
        """Validate if a URL is properly formatted."""
        if not url or not isinstance(url, str):
            return False
        try:
            from urllib.parse import urlparse
            result = urlparse(url)
            return all([result.scheme, result.netloc])
        except Exception:
            return False
    
    def _retry_operation(self, operation, max_retries=3, delay=1.0, operation_name="operation"):
        """
        Retry an operation with exponential backoff.
        
        Args:
            operation: Function to retry
            max_retries: Maximum number of retry attempts
            delay: Initial delay between retries in seconds
            operation_name: Name of the operation for logging
            
        Returns:
            Result of the operation or None if all retries fail
        """
        import time
        for attempt in range(max_retries):
            try:
                return operation()
            except (ConnectionError, TimeoutError) as e:
                if attempt < max_retries - 1:
                    wait_time = delay * (2 ** attempt)  # Exponential backoff
                    self._log('warn', f'{operation_name} failed (attempt {attempt + 1}/{max_retries}), retrying in {wait_time:.1f}s: {e}')
                    time.sleep(wait_time)
                else:
                    self._log('err', f'{operation_name} failed after {max_retries} attempts: {e}')
                    return None
            except Exception as e:
                self._log('err', f'{operation_name} failed with unexpected error: {e}')
                return None
    
    def _validate_document_integrity(self, file_path: str) -> Dict:
        """
        Validate document integrity and metadata.
        
        Args:
            file_path: Path to the document file
            
        Returns:
            Dictionary with validation results
        """
        validation = {
            "valid": True,
            "file_size": 0,
            "file_type": "unknown",
            "errors": [],
            "warnings": []
        }
        
        try:
            if not os.path.exists(file_path):
                validation["valid"] = False
                validation["errors"].append("File does not exist")
                return validation
            
            validation["file_size"] = os.path.getsize(file_path)
            
            # Check file size limits (GEM limit is 10MB)
            if validation["file_size"] > 10 * 1024 * 1024:  # 10MB
                validation["warnings"].append(f"File size ({validation['file_size'] / 1024 / 1024:.1f}MB) exceeds GEM limit of 10MB")
            
            # Check file type
            file_ext = os.path.splitext(file_path)[1].lower()
            validation["file_type"] = file_ext[1:] if file_ext else "unknown"
            
            if file_ext not in ['.pdf', '.doc', '.docx']:
                validation["warnings"].append(f"File type '{validation['file_type']}' may not be accepted by GEM")
            
            # PDF-specific validation
            if file_ext == '.pdf':
                try:
                    import pypdf
                    with open(file_path, 'rb') as f:
                        reader = pypdf.PdfReader(f)
                        page_count = len(reader.pages)
                        
                        # Check page limit (GEM limit is 100 pages)
                        if page_count > 100:
                            validation["warnings"].append(f"PDF has {page_count} pages, exceeding GEM limit of 100 pages")
                        
                        # Check if PDF is encrypted
                        if reader.is_encrypted:
                            validation["errors"].append("PDF is encrypted and cannot be processed")
                            validation["valid"] = False
                        
                        # Check for corrupted pages
                        corrupted_pages = []
                        for i, page in enumerate(reader.pages):
                            try:
                                page.extract_text()
                            except Exception:
                                corrupted_pages.append(i + 1)
                        
                        if corrupted_pages:
                            validation["warnings"].append(f"Pages {corrupted_pages} may be corrupted")
                
                except Exception as e:
                    validation["errors"].append(f"PDF validation failed: {str(e)}")
                    validation["valid"] = False
            
        except Exception as e:
            validation["valid"] = False
            validation["errors"].append(f"Validation error: {str(e)}")
        
        return validation
    
    def _validate_copied_documents(self):
        """
        Validate all copied documents for GEM compliance.
        Stores validation results and logs any issues.
        """
        self.validation_results = {}
        validation_issues = []
        
        for doc_name, doc_info in self.matched_documents.items():
            doc_path = doc_info.get('path') if isinstance(doc_info, dict) else doc_info
            if not doc_path or not os.path.exists(doc_path):
                continue
            
            validation = self._validate_document_integrity(doc_path)
            self.validation_results[doc_name] = validation
            
            if not validation["valid"]:
                validation_issues.append(f"{doc_name}: {', '.join(validation['errors'])}")
            elif validation["warnings"]:
                self._log('warn', f"{doc_name} warnings: {', '.join(validation['warnings'])}")
        
        if validation_issues:
            alert_document_issue(
                title="Document Validation Failed",
                message=f"Some documents failed validation:\n" + "\n".join(validation_issues),
                context={"validation_results": self.validation_results},
                severity=AlertSeverity.WARNING
            )
        else:
            self._log('ok', 'All documents validated successfully for GEM compliance')
    
    def _generate_validation_report(self, bid_no: str) -> str:
        """
        Generate a validation report for all copied documents.
        
        Args:
            bid_no: Tender bid number
            
        Returns:
            Path to the generated validation report file
        """
        report_path = os.path.join(self.filing_folder, "Document_Validation_Report.txt")
        
        try:
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write("Document Validation Report\n")
                f.write("=" * 50 + "\n\n")
                f.write(f"Tender: {bid_no}\n")
                f.write(f"Filing Folder: {self.filing_folder}\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
                
                f.write("GEM Compliance Validation\n")
                f.write("-" * 50 + "\n\n")
                
                for doc_name, validation in self.validation_results.items():
                    f.write(f"Document: {doc_name}\n")
                    f.write(f"  Status: {'✓ VALID' if validation['valid'] else '✗ INVALID'}\n")
                    f.write(f"  File Size: {validation['file_size'] / 1024 / 1024:.2f} MB\n")
                    f.write(f"  File Type: {validation['file_type']}\n")
                    
                    if validation['errors']:
                        f.write(f"  Errors:\n")
                        for error in validation['errors']:
                            f.write(f"    - {error}\n")
                    
                    if validation['warnings']:
                        f.write(f"  Warnings:\n")
                        for warning in validation['warnings']:
                            f.write(f"    - {warning}\n")
                    
                    f.write("\n")
                
                # Summary
                valid_count = sum(1 for v in self.validation_results.values() if v['valid'])
                total_count = len(self.validation_results)
                
                f.write("Summary\n")
                f.write("-" * 50 + "\n")
                f.write(f"Total Documents: {total_count}\n")
                f.write(f"Valid: {valid_count}\n")
                f.write(f"Invalid: {total_count - valid_count}\n")
                f.write(f"Compliance Rate: {valid_count / total_count * 100:.1f}%\n" if total_count > 0 else "Compliance Rate: N/A\n")
            
            self._log('ok', f'Generated validation report: {report_path}')
            return report_path
            
        except Exception as e:
            self._log('err', f'Failed to generate validation report: {e}')
            return ""
        
    def _log(self, level: str, message: str):
        """Internal logging method."""
        self.log_fn(level, message)

    def _save_step_logs_and_timings(self, step_timings: Dict = None, step_logs: Dict = None) -> str:
        """
        Save per-step logs and timing statistics to a dedicated 'step_logs' subfolder
        in the filing folder for offline debugging and test verification.
        """
        if not self.filing_folder or not os.path.exists(self.filing_folder):
            return ""
        
        step_logs_dir = os.path.join(self.filing_folder, "step_logs")
        try:
            os.makedirs(step_logs_dir, exist_ok=True)
        except Exception as e:
            self._log('warn', f"Failed to create step_logs directory: {e}")
            return ""
        
        # Save step_timings.json
        timings_path = os.path.join(step_logs_dir, "step_timings.json")
        try:
            with open(timings_path, 'w', encoding='utf-8') as f:
                json.dump(step_timings or self.processing_stats, f, indent=2)
        except Exception as e:
            self._log('warn', f"Failed to save step_timings.json: {e}")

        # Save individual step logs if available
        if step_logs and isinstance(step_logs, dict):
            for step_key, lines in step_logs.items():
                clean_name = re.sub(r'[^a-zA-Z0-9_]', '_', str(step_key)).lower()
                log_file = os.path.join(step_logs_dir, f"{clean_name}.log")
                try:
                    with open(log_file, 'w', encoding='utf-8') as f:
                        if isinstance(lines, list):
                            f.write("\n".join(lines))
                        else:
                            f.write(str(lines))
                except Exception as e:
                    self._log('warn', f"Failed to save step log file '{clean_name}.log': {e}")
                    
        return step_logs_dir
    
    def start_filing_process(self, tender_record: Dict, firm_name: str = None, category: str = None) -> Dict:
        """
        Start the complete filing process for a tender.
        
        Args:
            tender_record: Tender data dictionary from database
            firm_name: Optional firm name to use for document matching
            
        Returns:
            Dict with process results:
            {
                'success': bool,
                'filing_folder': str,
                'checklist_file': str,
                'summary_file': str,
                'required_count': int,
                'matched_count': int,
                'missing_count': int,
                'error': str (if failed)
            }
        """
        try:
            self.required_documents = []
            self.matched_documents = {}
            self.missing_documents = []
            self.filing_folder = ""
            self._local_llm_used = False
            self.category = category or "General"
            self.emd_details = {}
            self.gem_requirements = []

            bid_no = tender_record.get('bid_no', '')
            if not bid_no:
                return {'success': False, 'error': 'No bid number in tender record'}
            
            self._log('info', f'Starting filing process for {bid_no}')
            if self.progress_cb:
                self.progress_cb(5, "Initializing filing process...")
            
            # Step 1: Download tender PDF if not present
            if self.progress_cb:
                self.progress_cb(10, "Step 1: Downloading/verifying tender PDF...")
            pdf_path = self._ensure_tender_pdf(tender_record)
            if not pdf_path:
                return {'success': False, 'error': 'Failed to obtain tender PDF'}
            
            # Step 2: Create filing folder
            if self.progress_cb:
                self.progress_cb(20, "Step 2 & 3: Creating folder and copying tender PDF...")
            self.filing_folder = self._create_filing_folder(bid_no, firm_name, tender_record)
            self._log('ok', f'Created filing folder: {self.filing_folder}')
            
            # Step 3: Copy tender PDF to filing folder
            try:
                self._copy_tender_pdf(pdf_path)
            except (FileNotFoundError, PermissionError, OSError) as copy_err:
                self._log('warn', f'Failed to copy tender PDF: {copy_err}')
                alert_filing_issue(
                    title="PDF Copy Failed",
                    message=f"Failed to copy tender PDF to filing folder: {str(copy_err)}",
                    context={"pdf_path": pdf_path, "filing_folder": self.filing_folder, "error": str(copy_err)},
                    severity=AlertSeverity.WARNING
                )
                # Continue with the process even if copy fails
            except Exception as copy_err:
                self._log('warn', f'Unexpected error copying tender PDF: {copy_err}')
                alert_filing_issue(
                    title="Unexpected PDF Copy Error",
                    message=f"Unexpected error copying tender PDF: {str(copy_err)}",
                    context={"pdf_path": pdf_path, "error": str(copy_err)},
                    severity=AlertSeverity.WARNING
                )
                # Continue with the process even if copy fails
            
            # Step 4: Scan tender PDF for embedded links, download to Additional_Tender_Documents
            if self.progress_cb:
                self.progress_cb(35, "Step 4: Scanning PDF for embedded document links...")
            try:
                self._download_embedded_links(pdf_path)
            except (ConnectionError, TimeoutError) as link_err:
                self._log('warn', f'Network error downloading embedded links: {link_err}')
                alert_network_issue(
                    title="Embedded Links Download Failed",
                    message=f"Network error while downloading embedded links from PDF: {str(link_err)}",
                    context={"pdf_path": pdf_path, "error": str(link_err)},
                    severity=AlertSeverity.WARNING
                )
                # Continue with the process even if embedded link download fails
            except Exception as link_err:
                self._log('warn', f'Failed to download embedded links: {link_err}')
                alert_document_issue(
                    title="Embedded Links Processing Error",
                    message=f"Failed to process embedded links: {str(link_err)}",
                    context={"pdf_path": pdf_path, "error": str(link_err)},
                    severity=AlertSeverity.WARNING
                )
                # Continue with the process even if embedded link download fails
            
            # Step 5: Sourcing/Extracting Text from all files in Additional_Tender_Documents + main PDF
            if self.progress_cb:
                self.progress_cb(45, "Step 5: Extracting and parsing text from all files...")
            self._log('info', 'Extracting and scanning text from all downloaded documents (PDF, Excel, CSV)...')
            files_text = {}
            main_pdf_text = self._extract_pdf_text(pdf_path)
            if main_pdf_text:
                files_text['Tender_Document.pdf'] = main_pdf_text
                self.gem_requirements = self._extract_gem_requirements(main_pdf_text)
            else:
                self.gem_requirements = []
            
            add_folder = os.path.join(self.filing_folder, 'Additional_Tender_Documents')
            add_files_text = self._extract_text_from_additional_files(add_folder)
            files_text.update(add_files_text)
            
            combined_text = "\n\n".join(files_text.values())
            
            # Step 6: Identify Item Category using LLM
            if self.progress_cb:
                self.progress_cb(60, "Step 6: Identifying tender item category...")
            self._log('info', 'Identifying tender item category...')
            self.category = self._identify_category_and_docs(combined_text, tender_record)
            self._log('ok', f"Identified category: {self.category}")
            
            # Step 7: Required Documents Extraction with source file mapping
            if self.progress_cb:
                self.progress_cb(72, "Step 7: Extracting required document list...")
            self._log('info', 'Extracting required documents from all documents...')
            raw_documents = []
            # Extract requirements from main tender PDF text
            if main_pdf_text:
                main_docs = self._extract_required_documents(main_pdf_text)
                for doc in main_docs:
                    doc['source_file'] = 'Tender_Document.pdf'
                    raw_documents.append(doc)

            # Extract requirements from additional files
            for filename, f_text in add_files_text.items():
                file_docs = self._extract_required_documents(f_text)
                for doc in file_docs:
                    doc['source_file'] = filename
                    raw_documents.append(doc)
            
            # Remove duplicates and merge source_files
            seen_names = {}
            self.required_documents = []
            for doc in raw_documents:
                name_lower = doc['name'].lower()
                if name_lower not in seen_names:
                    seen_names[name_lower] = doc
                    self.required_documents.append(doc)
                else:
                    existing = seen_names[name_lower]
                    if doc['source_file'] not in existing.get('source_file', ''):
                        existing['source_file'] += f", {doc['source_file']}"
            
            self._log('ok', f'Found {len(self.required_documents)} required document types across all files')
            
            # Step 7.5: Resolve generic ATC document placeholders ("Additional Doc 1", "Additional Doc 2", etc.) using downloaded ATC documents
            self._log('info', 'Resolving generic ATC document definitions from downloaded documents...')
            self.required_documents = self._resolve_additional_doc_definitions(self.required_documents, add_files_text, main_pdf_text)

            # Alert if no required documents found
            if len(self.required_documents) == 0:
                alert_document_issue(
                    title="No Required Documents Found",
                    message="Failed to extract any required documents from tender PDF and additional documents",
                    context={"pdf_path": pdf_path, "tender_id": bid_no},
                    severity=AlertSeverity.WARNING
                )
            
            # Step 8: Extract EMD / Security Deposit details
            if self.progress_cb:
                self.progress_cb(80, "Step 8: Extracting EMD and security deposit requirements...")
            self._log('info', 'Extracting EMD and security requirements...')
            self.emd_details = self._extract_emd_details_llm(combined_text)
            emd_source = []
            for fname, text in files_text.items():
                if "emd" in text.lower() or "earnest money" in text.lower() or "performance security" in text.lower():
                    emd_source.append(fname)
            self.emd_details['source_file'] = ", ".join(emd_source) if emd_source else "Tender_Document.pdf"
            
            # Step 9: Match standard documents and category-specific documents with AI enhancement
            if self.progress_cb:
                self.progress_cb(85, "Step 9: Matching required documents with firm documents...")
            self._log('info', 'Matching required documents with firm documents...')
            active_firm = firm_name or tender_record.get('matched_firm') or "General"
            firm_documents = self._get_firm_documents(active_firm)
            
            # Use AI-enhanced matching for better accuracy
            self._log('info', 'Using AI-powered similarity scoring for document matching...')
            enhanced_matches = self._ai_enhanced_document_matching(self.required_documents, firm_documents)
            
            # Fall back to original matching if AI matching doesn't find enough matches
            if len(enhanced_matches) < len(self.required_documents) * 0.5:
                self._log('info', 'AI matching found limited results, using fallback matching...')
                self._match_documents(firm_documents)
                # Convert dict format to simple paths for compatibility
                simple_matches = {}
                for doc_name, doc_info in self.matched_documents.items():
                    if isinstance(doc_info, dict):
                        simple_matches[doc_name] = doc_info.get('path', doc_info)
                    else:
                        simple_matches[doc_name] = doc_info
                self.matched_documents = simple_matches
            else:
                self.matched_documents = enhanced_matches
                self._log('ok', f'AI-enhanced matching: {len(enhanced_matches)} documents matched with confidence scores')
            
            # Auto-pull category documents from settings memory
            self._auto_pull_category_documents(active_firm, self.category)
            
            # Step 10: Copy documents to filing folder with validation
            if self.progress_cb:
                self.progress_cb(92, "Step 10: Copying and validating matched documents...")
            self._log('info', 'Copying matched standard documents to filing folder...')
            self._copy_documents_to_folder()
            
            # Validate copied documents
            self._log('info', 'Validating copied documents for GEM compliance...')
            self._validate_copied_documents()
            
            # Copy common firm documents (GST, PAN, etc.)
            self._log('info', 'Copying common firm documents (GST, PAN, etc.)...')
            self._copy_common_firm_documents(firm_documents)
            
            # Alert if many documents are missing
            missing_ratio = len(self.missing_documents) / len(self.required_documents) if self.required_documents else 0
            if missing_ratio > 0.5 and len(self.required_documents) > 3:
                alert_filing_issue(
                    title="High Missing Document Ratio",
                    message=f"{len(self.missing_documents)} out of {len(self.required_documents)} required documents are missing ({missing_ratio:.0%})",
                    context={
                        "tender_id": bid_no,
                        "missing_count": len(self.missing_documents),
                        "required_count": len(self.required_documents),
                        "missing_docs": [d['name'] for d in self.missing_documents]
                    },
                    severity=AlertSeverity.WARNING
                )
            
            # Step 11: Auto-generate pre-filled compliance declarations (MII, Land Border, Non-Blacklisting, Bidder Undertaking, Bid Securing)
            if self.progress_cb:
                self.progress_cb(95, "Step 11: Auto-generating compliance declarations...")
            self._log('info', 'Auto-generating compliance declarations for active firm...')
            self._generate_auto_declarations(active_firm, tender_record)

            # Collect missing category-specific documents to return
            missing_category_docs = []
            for doc in self.missing_documents:
                if self._is_category_specific_doc(doc['name']):
                    missing_category_docs.append(doc)
                    
            # Generate checklist, summary, and Excel checklist
            checklist_file = self._generate_checklist(bid_no, tender_record)
            summary_file = self._generate_summary(bid_no, tender_record)
            excel_checklist_file = self._generate_excel_checklist(bid_no, tender_record)
            
            # Generate validation report
            validation_file = self._generate_validation_report(bid_no)
            
            # Export step logs & timing statistics to step_logs directory
            step_logs_dir = self._save_step_logs_and_timings()

            # Pre-filing readiness score audit
            readiness = self._calculate_filing_readiness()
            self._log('ok', f"Filing process completed! Readiness Score: {readiness['score']}% ({readiness['status_label']})")
            if self.progress_cb:
                self.progress_cb(100, f"Filing completed! Readiness Score: {readiness['score']}%")
            
            return {
                'success': True,
                'filing_folder': self.filing_folder,
                'checklist_file': checklist_file,
                'summary_file': summary_file,
                'excel_checklist_file': excel_checklist_file,
                'validation_file': validation_file,
                'step_logs_dir': step_logs_dir,
                'readiness': readiness,
                'required_count': len(self.required_documents),
                'matched_count': len(self.matched_documents),
                'missing_count': len(self.missing_documents),
                'missing_category_docs': missing_category_docs,
                'firm_name': active_firm,
                'category': self.category,
                'tender_record': tender_record,
                'required_documents': self.required_documents,
                'gem_requirements': self.gem_requirements,
                'validation_results': self.validation_results,
                'processing_stats': self.processing_stats
            }
        except FileNotFoundError as e:
            error_msg = f'File not found during filing: {str(e)}'
            self._log('err', error_msg)
            alert_filing_issue(
                title="File Not Found",
                message=error_msg,
                context={"error": str(e), "tender_id": bid_no},
                severity=AlertSeverity.ERROR
            )
            return {'success': False, 'error': error_msg}
        except PermissionError as e:
            error_msg = f'Permission denied during filing: {str(e)}'
            self._log('err', error_msg)
            alert_filing_issue(
                title="Permission Denied",
                message=error_msg,
                context={"error": str(e), "tender_id": bid_no},
                severity=AlertSeverity.ERROR
            )
            return {'success': False, 'error': error_msg}
        except (ConnectionError, TimeoutError) as e:
            error_msg = f'Network error during filing: {str(e)}'
            self._log('err', error_msg)
            alert_network_issue(
                title="Network Error in Filing",
                message=error_msg,
                context={"error": str(e), "tender_id": bid_no},
                severity=AlertSeverity.ERROR
            )
            return {'success': False, 'error': error_msg}
        except Exception as e:
            error_msg = f'Filing process failed: {str(e)}'
            self._log('err', error_msg)
            alert_filing_issue(
                title="Unexpected Filing Error",
                message=error_msg,
                context={"error": str(e), "tender_id": bid_no},
                severity=AlertSeverity.ERROR
            )
            return {'success': False, 'error': error_msg}
        finally:
            # Filing may use local LLM analysis for category and document
            # extraction.  Eject it once this job ends so its RAM is released.
            try:
                settings = db.load_settings()
                if (self._local_llm_used and
                        settings.get('llm_provider') == 'Local LLM (LM Studio / Ollama)'):
                    llm.unload_local_models(
                        settings.get('llm_base_url', ''),
                        settings.get('llm_api_key', ''),
                    )
                    self._log('info', 'Local LLM unloaded after filing completion.')
            except Exception as unload_error:
                self._log('warn', f'Could not unload local LLM: {unload_error}')

    def _extract_text_from_additional_files(self, folder_path: str) -> Dict[str, str]:
        """Extract text from all PDF, Excel, and CSV files in the folder."""
        files_text = {}
        if not os.path.exists(folder_path):
            return files_text

        for f in os.listdir(folder_path):
            f_path = os.path.join(folder_path, f)
            if not os.path.isfile(f_path):
                continue

            f_lower = f.lower()
            if f_lower.endswith('.pdf'):
                try:
                    with open(f_path, 'rb') as pdf_file:
                        pdf_bytes = pdf_file.read()
                    text = pdf_extractor.extract_text(pdf_bytes)
                    if text:
                        files_text[f] = text
                except Exception as e:
                    self._log('warn', f"Failed to extract text from additional PDF '{f}': {e}")
            elif f_lower.endswith(('.xlsx', '.xls')):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(f_path, read_only=True, data_only=True)
                    text_parts = []
                    for sheet in wb.worksheets:
                        text_parts.append(f"--- Sheet: {sheet.title} ---")
                        for row in sheet.iter_rows(values_only=True):
                            row_str = " | ".join(str(cell) for cell in row if cell is not None)
                            if row_str.strip():
                                text_parts.append(row_str)
                    files_text[f] = "\n".join(text_parts)
                except Exception as e:
                    self._log('warn', f"Failed to extract text from additional Excel '{f}': {e}")
            elif f_lower.endswith('.csv'):
                try:
                    with open(f_path, 'r', encoding='utf-8', errors='ignore') as csv_file:
                        files_text[f] = f"--- CSV File: {f} ---\n" + csv_file.read()
                except Exception as e:
                    self._log('warn', f"Failed to extract text from additional CSV '{f}': {e}")

        return files_text

    def _identify_category_and_docs(self, combined_text: str, tender_record: Dict) -> str:
        """Identify the item category of the tender using the LLM."""
        items_str = tender_record.get('items', '') or tender_record.get('category', '') or ""
        excerpt = combined_text[:4000]
        
        prompt = f"""
Analyze the GeM tender details:
Items/Category: "{items_str}"

Text Excerpt:
{excerpt}

Identify a single high-level product/service category for this tender (e.g. "Cables", "IT Hardware", "Medical Equipment", "Furniture", "Office Supplies", "General Services", etc.).
Return ONLY the category name in 1-3 words (e.g. "Cables" or "IT Hardware"). No other text, no markdown.
"""
        try:
            settings = db.load_settings()
            provider = settings.get('llm_provider', 'local')
            api_key = settings.get('llm_api_key', '')
            base_url = settings.get('llm_base_url', '')
            model = settings.get('llm_model', '')
            
            if provider == 'Local LLM (LM Studio / Ollama)':
                self._local_llm_used = True
            res_str = llm.call_llm(prompt, provider, api_key, base_url, model, response_json=False)
            category = res_str.strip().strip('"').strip("'")
            if "</think>" in category:
                category = category.split("</think>")[-1].strip()
            
            category = category.replace("\n", " ")
            category = re.sub(r'\s+', ' ', category)
            if len(category) > 40:
                category = "General"
            return category
        except Exception as e:
            self._log('warn', f"Failed to identify category via LLM: {e}")
            return "General"

    def _extract_emd_details_llm(self, combined_text: str) -> Dict:
        """Call LLM to extract EMD/Security deposit details from combined text."""
        excerpt = combined_text[:12000]
        prompt = f"""
Analyze the following text from a GeM tender and extract Earnest Money Deposit (EMD) or Security Deposit / Performance Bank Guarantee (PBG) details.

TEXT:
{excerpt}

Return ONLY a JSON object (no markdown, no backticks, no thinking block) with the following structure:
{{
  "emd_required": true or false,
  "emd_amount": "amount in Rs. or 'Exempted' or 'N/A'",
  "emd_exemption_allowed": true or false,
  "pbg_required": true or false,
  "pbg_percent": "percentage or amount or 'N/A'",
  "details_summary": "Brief summary of EMD/PBG instructions and bank details if any"
}}
"""
        try:
            settings = db.load_settings()
            provider = settings.get('llm_provider', 'local')
            api_key = settings.get('llm_api_key', '')
            base_url = settings.get('llm_base_url', '')
            model = settings.get('llm_model', '')
            
            if provider == 'Local LLM (LM Studio / Ollama)':
                self._local_llm_used = True
            res_str = llm.call_llm(prompt, provider, api_key, base_url, model, response_json=True)
            if "</think>" in res_str:
                res_str = res_str.split("</think>")[-1].strip()
                
            cleaned_json = llm.clean_json_response(res_str)
            return json.loads(cleaned_json)
        except Exception as e:
            self._log('warn', f"Failed to parse EMD details via LLM: {e}")
            return {
                "emd_required": False,
                "emd_amount": "N/A",
                "emd_exemption_allowed": False,
                "pbg_required": False,
                "pbg_percent": "N/A",
                "details_summary": "Failed to extract EMD details."
            }

    def _is_category_specific_doc(self, doc_name: str) -> bool:
        """Check if a document is category-specific."""
        standard_names = ["gst", "pan", "msme", "itr", "balance sheet", "turnover certificate", "shareholder pattern", "experience certificate"]
        doc_lower = doc_name.lower()
        return not any(std in doc_lower for std in standard_names)

    def _auto_pull_category_documents(self, firm_name: str, category: str):
        """Auto-pull category specific documents from last remembered locations."""
        settings = db.load_settings()
        firm_cat_docs = settings.get("firm_category_docs", {})
        saved_docs = firm_cat_docs.get(firm_name, {}).get(category, {})

        still_missing = []
        for doc in self.missing_documents:
            doc_name = doc['name']
            if self._is_category_specific_doc(doc_name):
                saved_path = saved_docs.get(doc_name)
                if saved_path and os.path.exists(saved_path):
                    dest_folder = os.path.join(self.filing_folder, "Category_Specific_Documents")
                    os.makedirs(dest_folder, exist_ok=True)
                    filename = os.path.basename(saved_path)
                    dest_path = os.path.join(dest_folder, filename)
                    
                    try:
                        shutil.copy2(saved_path, dest_path)
                        self.matched_documents[doc_name] = {
                            'path': dest_path,
                            'source': 'auto-pull (remembered)',
                            'document': doc
                        }
                        self._log('ok', f"[Auto-pull] Found and copied remembered category document '{doc_name}': {filename}")
                    except Exception as e:
                        self._log('warn', f"Failed to auto-pull remembered document '{doc_name}': {e}")
                        still_missing.append(doc)
                else:
                    still_missing.append(doc)
            else:
                still_missing.append(doc)
                
        self.missing_documents = still_missing

    def update_filing_results_with_new_matches(self, bid_no: str, tender_record: Dict, new_matches: Dict):
        """
        Merge new manually selected category matches and regenerate checklist and summary.
        """
        active_firm = tender_record.get('matched_firm') or "General"
        firm_documents = self._get_firm_documents(active_firm)
        
        self._match_documents(firm_documents)
        self._auto_pull_category_documents(active_firm, self.category)
        
        for doc_name, dest_path in new_matches.items():
            req_doc = next((d for d in self.required_documents if d['name'] == doc_name), None)
            if not req_doc:
                req_doc = {'name': doc_name, 'description': '', 'category': 'Technical', 'required': True}
                
            self.matched_documents[doc_name] = {
                'path': dest_path,
                'source': 'manual selection',
                'document': req_doc
            }
            self.missing_documents = [d for d in self.missing_documents if d['name'] != doc_name]
            
        self._generate_checklist(bid_no, tender_record)
        self._generate_summary(bid_no, tender_record)
        
    def _verify_pdf_bid_no(self, pdf_path: str, expected_bid_no: str) -> bool:
        """Verifies that the PDF file actually contains the expected bid number."""
        if not pdf_path or not os.path.exists(pdf_path):
            return False
        try:
            with open(pdf_path, 'rb') as f:
                text = pdf_extractor.extract_text(f.read())
            if not text:
                return False
            clean_text = " ".join(text.split()).lower()
            clean_bid = expected_bid_no.strip().lower()
            normalized_bid = clean_bid.replace("/", "").replace("_", "")
            normalized_text = clean_text.replace("/", "").replace("_", "")
            if clean_bid in clean_text or normalized_bid in normalized_text:
                return True
            return False
        except Exception:
            # Fallback to True to avoid false negatives on unreadable PDFs
            return True

    def _ensure_tender_pdf(self, tender_record: Dict) -> Optional[str]:
        """
        Ensure tender PDF is available locally.
        Download if not present.
        
        Args:
            tender_record: Tender data dictionary
            
        Returns:
            Path to PDF file or None if failed
        """
        bid_no = tender_record.get('bid_no', '')
        bid_url = tender_record.get('bid_url', '')
        existing_pdf = tender_record.get('pdf_path', '')
        
        # Check if PDF already exists locally
        if existing_pdf and os.path.exists(existing_pdf):
            if self._verify_pdf_bid_no(existing_pdf, bid_no):
                self._log('info', f'Using existing PDF: {existing_pdf}')
                return existing_pdf
            else:
                self._log('warn', f'Existing PDF {existing_pdf} does not match expected bid {bid_no}. Deleting and re-downloading.')
                try:
                    os.remove(existing_pdf)
                except Exception:
                    pass
        
        # Download PDF
        self._log('info', f'Downloading PDF for {bid_no}...')
        
        # Get download directory from settings
        settings = db.load_settings()
        download_dir = settings.get('pdf_save_folder', '')
        if not download_dir:
            download_dir = os.path.join(os.path.expanduser('~'), 'Documents', 'TenderPDFs')
        
        os.makedirs(download_dir, exist_ok=True)
        
        try:
            # Try to download using bid_url first
            if bid_url:
                # Bypass constructed fallback URL
                is_fallback_url = False
                id_m = re.search(r"([\dXx]+)$", bid_no)
                if id_m:
                    suffix = id_m.group(1)
                    if bid_url.rstrip("/").endswith(f"showbidDocument/{suffix}"):
                        is_fallback_url = True
                
                if is_fallback_url:
                    self._log('info', f'Bypassing constructed fallback URL for {bid_no}')
                else:
                    self._log('info', f'Attempting download using URL: {bid_url}')
                    pdf_path = scraper.download_tender_pdf(
                        bid_url, 
                        download_dir, 
                        log_fn=self._log, 
                        headless=True
                    )
                    if pdf_path:
                        if self._verify_pdf_bid_no(pdf_path, bid_no):
                            db.upsert_tender_field(bid_no, 'pdf_path', pdf_path)
                            return pdf_path
                        else:
                            self._log('warn', f'Downloaded PDF from URL does not match {bid_no}. Deleting and trying search fallback.')
                            try:
                                os.remove(pdf_path)
                            except Exception:
                                pass
            
            # Fallback to bid_no search
            if bid_no:
                self._log('info', f'Attempting search download for {bid_no}...')
                pdf_path = scraper.download_tender_pdf(
                    bid_no, 
                    download_dir, 
                    log_fn=self._log, 
                    headless=True
                )
                if pdf_path:
                    if self._verify_pdf_bid_no(pdf_path, bid_no):
                        db.upsert_tender_field(bid_no, 'pdf_path', pdf_path)
                        return pdf_path
                    else:
                        self._log('err', f'Downloaded PDF from search does not match {bid_no}. Deleting.')
                        try:
                            os.remove(pdf_path)
                        except Exception:
                            pass
                    
        except Exception as e:
            self._log('err', f'PDF download failed: {e}')
        
        return None
    
    def _extract_pdf_text(self, pdf_path: str) -> Optional[str]:
        """
        Extract text from PDF file.
        
        Args:
            pdf_path: Path to PDF file
            
        Returns:
            Extracted text or None if failed
        """
        try:
            with open(pdf_path, 'rb') as f:
                pdf_bytes = f.read()
            
            pdf_text = pdf_extractor.extract_text(pdf_bytes)
            if pdf_text:
                self._log('ok', f'Extracted {len(pdf_text)} characters from PDF')
                return pdf_text
            else:
                self._log('warn', 'PDF text extraction returned empty result')
                return None
                
        except Exception as e:
            self._log('err', f'PDF text extraction failed: {e}')
            return None
    
    def _extract_required_documents(self, pdf_text: str) -> List[Dict]:
        """
        Extract required documents from PDF text using hybrid regex/LLM approach.
        
        Args:
            pdf_text: Extracted text from PDF
            
        Returns:
            List of document requirement dictionaries:
            [{
                'name': str,
                'description': str,
                'category': str,
                'required': bool,
                'source': 'regex' | 'llm'
            }]
        """
        documents = []
        
        # Step 1: Try regex patterns for common document types
        regex_docs = self._extract_documents_regex(pdf_text)
        documents.extend(regex_docs)
        
        # Step 2: Use an enabled LLM to supplement the deterministic rules.
        # Tender wording is not standardized, so a count-based fallback can
        # miss a requirement simply because the regex already found three
        # unrelated documents.
        provider = db.load_settings().get('llm_provider', 'Disabled')
        if provider not in ('', 'Disabled', None):
            self._log('info', 'Supplementing regex results with LLM analysis...')
            llm_docs = self._extract_documents_llm(pdf_text)
            documents.extend(llm_docs)
        
        # Remove duplicates based on name
        seen_names = set()
        unique_docs = []
        for doc in documents:
            doc_name_lower = doc['name'].lower()
            if doc_name_lower not in seen_names:
                seen_names.add(doc_name_lower)
                unique_docs.append(doc)
        
        return unique_docs
    
    def _extract_documents_regex(self, pdf_text: str) -> List[Dict]:
        """
        Extract document requirements using regex patterns.
        
        Args:
            pdf_text: PDF text
            
        Returns:
            List of document dictionaries
        """
        documents = []
        text_lower = pdf_text.lower()
        
        # Common document patterns with descriptions
        document_patterns = [
            (r'gst\s*(certificate|registration|registration\s*certificate)', 'GST Certificate', 'Financial'),
            (r'pan\s*(card|number)', 'PAN Card', 'Financial'),
            (r'aadh?a?r\s*(card|number)?', 'Aadhaar Card', 'Legal'),
            (r'msme\s*(certificate|registration|udyam\s*registration)', 'MSME Certificate', 'Legal'),
            (r'income\s*tax\s*return|itr', 'Income Tax Return (ITR)', 'Financial'),
            (r'balance\s*sheet', 'Balance Sheet', 'Financial'),
            (r'turnover\s*certificate', 'Turnover Certificate', 'Financial'),
            (r'shareholder\s*(pattern|details|certificate)', 'Shareholder Pattern', 'Legal'),
            (r'experience\s*certificate', 'Experience Certificate', 'Technical'),
            (r'oem\s*(authorization|certificate)', 'OEM Authorization', 'Technical'),
            (r'mii\s*(declaration|certificate|compliance)', 'MII Declaration', 'Technical'),
            (r'technical\s*specifications', 'Technical Specifications', 'Technical'),
            (r'annexure', 'Annexure', 'Technical'),
            (r'bid\s*security\s*deposit|emd', 'EMD Document', 'Financial'),
            (r'epbg\s*(document|certificate)', 'ePBG Document', 'Financial'),
        ]
        
        for pattern, name, category in document_patterns:
            if re.search(pattern, text_lower):
                documents.append({
                    'name': name,
                    'description': f'{name} as specified in tender document',
                    'category': category,
                    'required': True,
                    'source': 'regex'
                })

        # Eligibility clauses frequently omit words such as "certificate" or
        # "document".  Extract evidence requirements from the criterion
        # itself, regardless of whether the tender puts the number before or
        # after "experience"/"turnover", uses number words, or writes
        # "turn over" as two words.
        number_words = {
            'one': 1, 'two': 2, 'three': 3, 'four': 4, 'five': 5,
            'six': 6, 'seven': 7, 'eight': 8, 'nine': 9, 'ten': 10,
        }
        number_pattern = r'\d+|one|two|three|four|five|six|seven|eight|nine|ten'

        def normalized_number(value: str) -> str:
            return str(number_words.get(value, int(value))) if value.isdigit() else str(number_words[value])

        experience_patterns = [
            re.compile(
                rf'\b(?P<years>{number_pattern})\s*(?:-\s*)?(?:years?|yrs?)\s*(?:of\s+)?'
                r'(?:(?:past|similar)\s+)?experience(?:\s+(?:of|in|with)\s+(?P<scope>[^,;\n.]+?))?'
                r'(?=\s*(?:is|are|must|shall|will|and|,|;|\.|\n|$))'
            ),
            re.compile(
                rf'\b(?:past\s+|similar\s+)?experience\s*(?:required)?\s*(?::|-|of|for|in)\s*'
                rf'(?P<years>{number_pattern})\s*(?:-\s*)?(?:years?|yrs?)'
                r'(?:\s+(?:of|in|with)\s+(?P<scope>[^,;\n.]+?))?'
                r'(?=\s*(?:is|are|must|shall|will|and|,|;|\.|\n|$))'
            ),
        ]
        for pattern in experience_patterns:
            experience_match = pattern.search(text_lower)
            if not experience_match:
                continue
            years = normalized_number(experience_match.group('years'))
            scope = (experience_match.groupdict().get('scope') or '').strip()
            scope_description = f' of {scope}' if scope else ''
            documents.append({
                'name': f'Experience Proof - {years} Years',
                'description': (
                    f'Proof of {years} years of experience{scope_description} '
                    'as specified in tender document'
                ),
                'category': 'Technical',
                'required': True,
                'source': 'regex'
            })
            break

        turnover_patterns = [
            re.compile(
                rf'\b(?:last|previous|past)\s+(?P<period>{number_pattern})\s*(?:-\s*)?'
                r'(?:financial\s+)?years?\s+(?:turn\s*over|turnover)\b'
            ),
            re.compile(
                rf'\b(?:turn\s*over|turnover)\b[^.\n;]{{0,80}}?\b'
                rf'(?:last|previous|past)\s+(?P<period>{number_pattern})\s*(?:-\s*)?(?:financial\s+)?years?\b'
            ),
        ]
        for pattern in turnover_patterns:
            turnover_match = pattern.search(text_lower)
            if not turnover_match:
                continue
            period = normalized_number(turnover_match.group('period'))
            documents.append({
                'name': f'Turnover Proof - Last {period} Years',
                'description': (
                    f'Turnover proof for the last {period} years as specified '
                    'in tender document'
                ),
                'category': 'Financial',
                'required': True,
                'source': 'regex'
            })
            break
        
        # Check for year-specific requirements
        year_patterns = [
            (r'itr.*(\d+)', 'ITR - Year {}', 'Financial'),
            (r'balance\s*sheet.*(\d+)', 'Balance Sheet - Year {}', 'Financial'),
        ]
        
        for pattern, name_template, category in year_patterns:
            matches = re.finditer(pattern, text_lower)
            for match in matches:
                year = match.group(1)
                documents.append({
                    'name': name_template.format(year),
                    'description': f'{name_template.format(year)} as specified in tender',
                    'category': category,
                    'required': True,
                    'source': 'regex'
                })
        
        return documents
    
    def _extract_documents_llm(self, pdf_text: str) -> List[Dict]:
        """
        Extract document requirements using LLM analysis.
        
        Args:
            pdf_text: PDF text
            
        Returns:
            List of document dictionaries
        """
        try:
            settings = db.load_settings()
            provider = settings.get('llm_provider', 'Disabled')
            if provider == 'Disabled':
                self._log('warn', 'LLM provider disabled, skipping LLM document extraction')
                return []

            if provider == 'Local LLM (LM Studio / Ollama)':
                self._local_llm_used = True
            
            # Keep input small — thinking models need tokens for reasoning + output.
            # 4000 chars is enough to capture eligibility/document clauses near the
            # top of the tender text without exhausting the token budget.
            truncated_text = pdf_text[:4000] if len(pdf_text) > 4000 else pdf_text
            
            prompt = f"""Extract all document requirements from this tender document.
Capture explicit documents and evidence implied by eligibility clauses. For example, if the tender requires experience, turnover, or a tax registration, return the supporting proof as a required document.

Tender text:
{truncated_text}

Return a JSON array of documents with this exact format:
[
  {{
    "name": "Document Name",
    "description": "Brief description",
    "category": "Financial|Legal|Technical",
    "required": true
  }}
]

Return ONLY the JSON array, no other text."""
            
            response_text = llm.call_llm(prompt, provider, 
                                        settings.get('llm_api_key', ''),
                                        settings.get('llm_base_url', ''),
                                        settings.get('llm_model', ''),
                                        response_json=True)
            
            # Strip thinking/reasoning block if present before any JSON cleaning
            if '</think>' in response_text:
                response_text = response_text.split('</think>')[-1].strip()
            
            cleaned_json = llm.clean_json_response(response_text)
            
            # Try normal parse first; fall back to partial-repair for truncated output
            try:
                parsed_docs = json.loads(cleaned_json)
            except json.JSONDecodeError:
                repaired = llm.repair_truncated_json_array(cleaned_json)
                parsed_docs = json.loads(repaired)
            
            documents = []
            for doc in parsed_docs:
                if isinstance(doc, dict):
                    documents.append({
                        'name': doc.get('name', 'Unknown Document'),
                        'description': doc.get('description', ''),
                        'category': doc.get('category', 'Technical'),
                        'required': doc.get('required', True),
                        'source': 'llm'
                    })
            
            self._log('ok', f'LLM extracted {len(documents)} document requirements')
            return documents
            
        except Exception as e:
            if "not configured" in str(e) or "disabled" in str(e).lower():
                self._log('info', f'LLM skipped: {e}')
            else:
                self._log('warn', f'LLM document extraction skipped: {e}')
            return []

    
    def _get_firm_documents(self, firm_name: str = None) -> Dict:
        """
        Get firm documents from settings.
        
        Args:
            firm_name: Optional firm name to filter by
            
        Returns:
            Dictionary of firm documents
        """
        settings = db.load_settings()
        firms = settings.get('firms', [])
        
        target_firm = None
        if firm_name:
            for firm in firms:
                if firm.get('name') == firm_name:
                    target_firm = firm
                    break
        
        # If no specific firm or not found, use first firm
        if not target_firm and firms:
            target_firm = firms[0]
        
        if target_firm:
            return target_firm.get('documents', {})
        
        return {}
    
    def _find_category_folder(self, filing_folder: str, category: str, base_dir: str = "") -> Optional[str]:
        """Find matching category subfolder (e.g. 'Cable', 'Cables', 'Cable Documents')."""
        if not category or category in ('General', 'N/A'):
            return None

        cat_clean = category.strip().lower()
        search_dirs = []
        if filing_folder:
            search_dirs.append(os.path.normpath(os.path.join(filing_folder, '..')))
            search_dirs.append(os.path.normpath(os.path.join(filing_folder, '..', '..')))
            search_dirs.append(os.path.normpath(os.path.join(filing_folder, '..', 'COMMON')))
        if base_dir:
            search_dirs.append(os.path.normpath(base_dir))
            search_dirs.append(os.path.normpath(os.path.join(base_dir, 'COMMON')))

        seen = set()
        # 1. Exact or plural match
        for p_dir in search_dirs:
            if not p_dir or p_dir in seen or not os.path.exists(p_dir) or not os.path.isdir(p_dir):
                continue
            seen.add(p_dir)
            try:
                for item in os.listdir(p_dir):
                    item_path = os.path.join(p_dir, item)
                    if os.path.isdir(item_path):
                        item_clean = item.strip().lower()
                        if (item_clean == cat_clean or
                            item_clean == cat_clean + 's' or
                            cat_clean == item_clean + 's'):
                            return item_path
            except Exception:
                pass

        # 2. Substring match
        seen.clear()
        for p_dir in search_dirs:
            if not p_dir or p_dir in seen or not os.path.exists(p_dir) or not os.path.isdir(p_dir):
                continue
            seen.add(p_dir)
            try:
                for item in os.listdir(p_dir):
                    item_path = os.path.join(p_dir, item)
                    if os.path.isdir(item_path) and item.upper() != "COMMON":
                        item_clean = item.strip().lower()
                        if cat_clean in item_clean or item_clean in cat_clean:
                            return item_path
            except Exception:
                pass

        return None

    def _match_documents(self, firm_documents: Dict):
        """
        Match required documents with available firm documents, COMMON folder files,
        and category-specific folder files.
        """
        self.matched_documents = {}
        self.missing_documents = []
        
        # Document type mapping for fuzzy matching
        type_mappings = {
            'gst': ['gst', 'gst certificate', 'gst registration'],
            'pan': ['pan', 'pan card'],
            'msme': ['msme', 'msme certificate', 'udyam'],
            'itr': ['itr', 'income tax', 'income tax return'],
            'bs': ['balance sheet', 'balance sheet'],
            'turnover': ['turnover', 'turnover certificate'],
            'experience': ['experience', 'experience proof', 'experience certificate'],
            'aadhaar': ['aadhaar', 'aadhar', 'aadhaar card', 'aadhar card'],
            'shareholder': ['shareholder', 'shareholder pattern'],
        }
        
        # Add COMMON folder files to available documents
        common_folder_path = os.path.join(self.filing_folder, '..', 'COMMON')
        common_documents = {}
        if os.path.exists(common_folder_path) and os.path.isdir(common_folder_path):
            for item in os.listdir(common_folder_path):
                item_path = os.path.join(common_folder_path, item)
                if os.path.isfile(item_path):
                    common_documents[item] = item_path
            if common_documents:
                self._log('info', f'Using {len(common_documents)} COMMON folder files for document matching')

        # Add Category subfolder files to available documents
        category_documents = {}
        category = getattr(self, 'category', '')
        if category and category != 'General':
            category_folder_path = self._find_category_folder(self.filing_folder, category)
            if category_folder_path and os.path.exists(category_folder_path) and os.path.isdir(category_folder_path):
                for item in os.listdir(category_folder_path):
                    item_path = os.path.join(category_folder_path, item)
                    if os.path.isfile(item_path):
                        category_documents[item] = item_path
                if category_documents:
                    self._log('info', f"Using {len(category_documents)} category-specific files from '{os.path.basename(category_folder_path)}' folder")

        # Merge firm documents, COMMON folder documents, and Category folder documents
        all_documents = {**firm_documents, **common_documents, **category_documents}
        
        for req_doc in self.required_documents:
            req_name_lower = req_doc['name'].lower()
            matched = False
            
            # Try exact matches first
            for doc_key, doc_path in all_documents.items():
                if doc_path and isinstance(doc_path, str) and os.path.exists(doc_path):
                    doc_key_lower = doc_key.lower()
                    if req_name_lower in doc_key_lower or doc_key_lower in req_name_lower:
                        if doc_key in common_documents:
                            source = 'COMMON'
                        elif doc_key in category_documents:
                            source = f'Category ({category})'
                        else:
                            source = doc_key
                        self.matched_documents[req_doc['name']] = {
                            'path': doc_path,
                            'source': source,
                            'document': req_doc
                        }
                        matched = True
                        break
            
            # Try fuzzy matches
            if not matched:
                for doc_key, doc_path in all_documents.items():
                    if doc_path and isinstance(doc_path, str) and os.path.exists(doc_path):
                        for map_type, keywords in type_mappings.items():
                            if any(kw in req_name_lower for kw in keywords) and map_type in doc_key.lower():
                                if doc_key in common_documents:
                                    source = 'COMMON'
                                elif doc_key in category_documents:
                                    source = f'Category ({category})'
                                else:
                                    source = doc_key
                                self.matched_documents[req_doc['name']] = {
                                    'path': doc_path,
                                    'source': source,
                                    'document': req_doc
                                }
                                matched = True
                                break
                    if matched:
                        break
            
            if not matched:
                self.missing_documents.append(req_doc)
    
    def _create_filing_folder(self, bid_no: str, firm_name: str = None, tender_record: Dict = None) -> str:
        """
        Create filing folder for the tender.
        
        Args:
            bid_no: Tender bid number
            firm_name: Optional firm name for folder naming
            tender_record: Tender data dictionary for extracting date and category
            
        Returns:
            Path to created folder
        """
        # Get base directory from settings or use Documents
        settings = db.load_settings()
        base_dir = settings.get('pdf_save_folder', '')
        if not base_dir:
            base_dir = os.path.join(os.path.expanduser('~'), 'Documents', 'TenderPDFs')
        
        # Create base folder structure: {base}\GEM TENDER x {firm_name} {financial_year}
        fy = financial_year(datetime.now())
        if firm_name:
            base_folder_name = f"GEM TENDER x {firm_name} {fy}"
        else:
            base_folder_name = f"GEM TENDER {fy}"
        
        # Sanitize base folder name
        base_folder_name = re.sub(r'[<>:"/\\|?*]', '_', base_folder_name)
        
        base_folder_path = os.path.join(base_dir, base_folder_name)
        os.makedirs(base_folder_path, exist_ok=True)
        
        # Look for COMMON subfolder for shared firm documents
        common_folder_path = os.path.join(base_folder_path, 'COMMON')
        common_files = []
        
        if os.path.exists(common_folder_path) and os.path.isdir(common_folder_path):
            for item in os.listdir(common_folder_path):
                item_path = os.path.join(common_folder_path, item)
                # Only copy files, not subdirectories
                if os.path.isfile(item_path):
                    common_files.append(item_path)
            
            if common_files:
                self._log('info', f'Found {len(common_files)} common file(s) in COMMON folder')
        
        # Create subfolder: {yyyymmdd} {gemid} {firmname} {category}
        if tender_record:
            # Extract date from tender record (only use end_date, do not fall back)
            raw_date = tender_record.get('end_date') or ''
            date_str = ''
            if raw_date:
                raw_date = str(raw_date).strip()
                # Parse common formats: dd-mm-yyyy, dd/mm/yyyy, yyyy-mm-dd, etc.
                for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y %I:%M %p", "%d-%m-%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f"):
                    try:
                        date_obj = datetime.strptime(raw_date, fmt)
                        date_str = date_obj.strftime('%Y%m%d')
                        break
                    except ValueError:
                        pass
                
                # If still not parsed, try parsing first 10 characters (the date portion) with basic formats
                if not date_str and len(raw_date) >= 10:
                    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
                        try:
                            date_obj = datetime.strptime(raw_date[:10], fmt)
                            date_str = date_obj.strftime('%Y%m%d')
                            break
                        except ValueError:
                            pass
                
            if not date_str:
                # Prompt the user to enter the end date
                try:
                    import tkinter as tk
                    from tkinter import simpledialog
                    
                    root = tk.Tk()
                    root.withdraw()
                    
                    user_date = simpledialog.askstring(
                        "End Date Required",
                        f"The tender '{bid_no}' has no valid end date.\n"
                        f"Please enter the tender end date (e.g. 14/07/2026 or 14-07-2026):",
                        parent=root
                    )
                    root.destroy()
                    
                    if user_date:
                        user_date = user_date.strip()
                        for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
                            try:
                                date_obj = datetime.strptime(user_date, fmt)
                                date_str = date_obj.strftime('%Y%m%d')
                                tender_record['end_date'] = user_date
                                db.upsert_tender(tender_record)
                                break
                            except ValueError:
                                pass
                except Exception:
                    pass
                
                if not date_str:
                    try:
                        import sys
                        # Check if stdin/stdout are terminals
                        if sys.stdin.isatty() and sys.stdout.isatty():
                            print(f"\n[PROMPT] End Date not available for tender '{bid_no}'.")
                            val = input("Please enter the end date (dd/mm/yyyy): ").strip()
                            for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
                                try:
                                    date_obj = datetime.strptime(val, fmt)
                                    date_str = date_obj.strftime('%Y%m%d')
                                    tender_record['end_date'] = val
                                    db.upsert_tender(tender_record)
                                    break
                                except ValueError:
                                    pass
                    except Exception:
                        pass

            if not date_str:
                raise ValueError("Tender end_date is mandatory and not available or valid.")
            
            # Extract GEM ID from bid_no (e.g., GEM/2026/B/7719019 -> GEM2026B7719019)
            gem_id = bid_no.replace('/', '').replace('-', '')
            
            # Get category (use category from tender record or default to "General")
            category = tender_record.get('category', 'General')
            if not category or category == 'N/A':
                category = 'General'
            
            # Apply category mapping from refine rules to get the standardized category
            try:
                from parser import map_category
                # Use items text for better category mapping
                items_text = tender_record.get('items', '') or tender_record.get('title', '')
                if items_text:
                    mapped_category = map_category(items_text, allow_llm=False)
                    if mapped_category and mapped_category != items_text.strip().title():
                        category = mapped_category
            except Exception as map_err:
                self._log('warn', f'Category mapping failed: {map_err}')
            
            # Sanitize category for folder name and truncate if too long
            category = re.sub(r'[<>:"/\\|?*]', '_', category)
            # Limit category to 15 characters to avoid path length issues
            if len(category) > 15:
                category = category[:15]
            
            # Build subfolder name
            if firm_name:
                # Sanitize firm name for folder name
                safe_firm_name = re.sub(r'[<>:"/\\|?*]', '_', firm_name)
                # Limit firm name to 20 characters
                if len(safe_firm_name) > 20:
                    safe_firm_name = safe_firm_name[:20]
                subfolder_name = f"{date_str} {gem_id} {safe_firm_name} {category}"
            else:
                subfolder_name = f"{date_str} {gem_id} {category}"
        else:
            # Fallback if no tender record
            date_str = datetime.now().strftime('%Y%m%d')
            gem_id = bid_no.replace('/', '').replace('-', '')
            if firm_name:
                safe_firm_name = re.sub(r'[<>:"/\\|?*]', '_', firm_name)
                subfolder_name = f"{date_str} {gem_id} {safe_firm_name}"
            else:
                subfolder_name = f"{date_str} {gem_id}"
        
        # Sanitize subfolder name
        subfolder_name = re.sub(r'[<>:"/\\|?*]', '_', subfolder_name)
        
        folder_path = os.path.join(base_folder_path, subfolder_name)
        
        # Validate path length before creating
        if len(folder_path) > 250:  # Windows MAX_PATH limit
            self._log('warn', f'Folder path too long ({len(folder_path)} chars), truncating further')
            # Truncate subfolder name more aggressively
            subfolder_name = subfolder_name[:100]
            folder_path = os.path.join(base_folder_path, subfolder_name)
        
        os.makedirs(folder_path, exist_ok=True)
        
        # Look for Category subfolder for category-specific documents
        category_files = []
        cat_folder = self._find_category_folder(folder_path, category, base_dir)
        if cat_folder and os.path.exists(cat_folder) and os.path.isdir(cat_folder):
            for item in os.listdir(cat_folder):
                item_path = os.path.join(cat_folder, item)
                if os.path.isfile(item_path):
                    category_files.append(item_path)
            if category_files:
                self._log('info', f"Found {len(category_files)} file(s) in Category folder '{os.path.basename(cat_folder)}'")

        # Copy common files from base folder to new subfolder
        if common_files:
            copied_count = 0
            for common_file in common_files:
                try:
                    filename = os.path.basename(common_file)
                    dest_path = os.path.join(folder_path, filename)
                    if not os.path.exists(dest_path):
                        shutil.copy2(common_file, dest_path)
                        copied_count += 1
                except Exception as copy_err:
                    self._log('warn', f'Failed to copy common file {os.path.basename(common_file)}: {copy_err}')
            
            if copied_count > 0:
                self._log('ok', f'Copied {copied_count} common file(s) to new folder')

        # Copy category-specific files from Category folder to new subfolder
        if category_files:
            copied_cat_count = 0
            for cat_file in category_files:
                try:
                    filename = os.path.basename(cat_file)
                    dest_path = os.path.join(folder_path, filename)
                    if not os.path.exists(dest_path):
                        shutil.copy2(cat_file, dest_path)
                        copied_cat_count += 1
                except Exception as copy_err:
                    self._log('warn', f'Failed to copy category file {os.path.basename(cat_file)}: {copy_err}')
            
            if copied_cat_count > 0:
                self._log('ok', f"Copied {copied_cat_count} category file(s) from '{os.path.basename(cat_folder)}' folder to new folder")

        return folder_path
    
    def _copy_documents_to_folder(self):
        """Copy matched documents to filing folder."""
        for doc_name, doc_info in self.matched_documents.items():
            try:
                # Handle both dict (AI matching) and string (original matching) formats
                if isinstance(doc_info, dict):
                    src_path = doc_info.get('path')
                else:
                    src_path = doc_info
                
                if not src_path or not os.path.exists(src_path):
                    self._log('warn', f'Source path not found for {doc_name}')
                    continue
                
                # Generate safe filename
                safe_name = re.sub(r'[<>:"/\\|?*]', '_', doc_name)
                ext = os.path.splitext(src_path)[1]
                dest_path = os.path.join(self.filing_folder, f"{safe_name}{ext}")
                
                # Handle duplicate filenames
                counter = 1
                while os.path.exists(dest_path):
                    dest_path = os.path.join(self.filing_folder, f"{safe_name}_{counter}{ext}")
                    counter += 1
                
                shutil.copy2(src_path, dest_path)
                self._log('ok', f'Copied {doc_name} to filing folder')
                    
            except Exception as e:
                self._log('err', f'Failed to copy {doc_name}: {e}')

    def _copy_tender_pdf(self, pdf_path: str):
        """Copy the source tender PDF into the filing folder for reference."""
        try:
            ext = os.path.splitext(pdf_path)[1] or '.pdf'
            dest_path = os.path.join(self.filing_folder, f'Tender_Document{ext}')
            shutil.copy2(pdf_path, dest_path)
            self._log('ok', 'Copied tender PDF to filing folder')
        except Exception as e:
            # The checklist can still be useful even if the source copy fails.
            self._log('warn', f'Failed to copy tender PDF: {e}')

    def _copy_common_firm_documents(self, firm_documents: Dict):
        """Copy all uploaded firm documents to a common subfolder."""
        if not firm_documents:
            return
            
        common_folder = os.path.join(self.filing_folder, "Common_Firm_Documents")
        try:
            os.makedirs(common_folder, exist_ok=True)
            for doc_key, doc_path in firm_documents.items():
                if doc_path and isinstance(doc_path, str) and os.path.exists(doc_path):
                    # Mapping from database key to readable name
                    name_mappings = {
                        "gst": "GST_Certificate",
                        "pan": "PAN_Card",
                        "msme": "MSME_Certificate",
                        "itr_1": "ITR_Year_1",
                        "itr_2": "ITR_Year_2",
                        "itr_3": "ITR_Year_3",
                        "bs_1": "Balance_Sheet_Year_1",
                        "bs_2": "Balance_Sheet_Year_2",
                        "bs_3": "Balance_Sheet_Year_3",
                        "turnover_cert": "Turnover_Certificate",
                        "shareholder": "Shareholder_Pattern"
                    }
                    display_name = name_mappings.get(doc_key, doc_key)
                    safe_name = re.sub(r'[<>:"/\\|?*]', '_', display_name)
                    ext = os.path.splitext(doc_path)[1]
                    dest_path = os.path.join(common_folder, f"{safe_name}{ext}")
                    
                    shutil.copy2(doc_path, dest_path)
                    self._log('ok', f"Copied firm document '{display_name}' to Common_Firm_Documents")
        except Exception as e:
            self._log('warn', f"Failed to copy common firm documents: {e}")
    
    def _download_embedded_links(self, pdf_path: str):
        """Find and download all embedded links (PDFs, Excel BOQs) from the tender PDF."""
        self._log('info', 'Scanning tender PDF for embedded links...')
        
        try:
            with open(pdf_path, 'rb') as f:
                pdf_bytes = f.read()
        except Exception as e:
            self._log('warn', f'Failed to read PDF for link extraction: {e}')
            return

        urls = set()

        # 1. Extract links via PyMuPDF (fitz) if available
        try:
            import fitz
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            for page in doc:
                for link in page.get_links():
                    uri = link.get("uri")
                    if uri and uri.strip().lower().startswith("http"):
                        urls.add(uri.strip())
            doc.close()
        except Exception:
            pass

        # 2. Extract links via pypdf (fallback if fitz is not installed)
        try:
            import pypdf
            import io
            reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
            for page in reader.pages:
                if "/Annots" in page:
                    annots = page["/Annots"]
                    if not isinstance(annots, list):
                        annots = annots.get_object()
                    for annot in annots:
                        annot_obj = annot.get_object()
                        if annot_obj.get("/Subtype") == "/Link":
                            action = annot_obj.get("/A")
                            if action:
                                action_obj = action.get_object()
                                if "/URI" in action_obj:
                                    uri = action_obj["/URI"]
                                    if uri:
                                        urls.add(str(uri).strip())
        except Exception as e:
            self._log('warn', f'pypdf link annotation extraction failed: {e}')

        # 3. Extract plain text URLs using regex from extracted text
        try:
            text = pdf_extractor.extract_text(pdf_bytes)
            text_urls = re.findall(r'https?://[^\s)\]"\',]+', text, re.IGNORECASE)
            for u in text_urls:
                if "gem.gov.in" in u.lower():
                    urls.add(u.strip())
        except Exception:
            pass

        if not urls:
            self._log('ok', 'No embedded links found in tender PDF.')
            return

        import urllib.parse

        def get_valid_document_url_and_filename(url_str):
            valid_exts = ('.pdf', '.xls', '.xlsx', '.doc', '.docx', '.zip', '.csv')
            clean_u = url_str.rstrip('.')
            if "gem.gov.in" not in clean_u.lower():
                return None, None
                
            parsed_url = urllib.parse.urlparse(clean_u)
            
            # 1. Try to extract filename from URL path first
            path = parsed_url.path
            path_lower = path.lower()
            for ext in valid_exts:
                if path_lower.endswith(ext):
                    filename = os.path.basename(path)
                    return clean_u, filename
                    
            # 2. Try to extract from query parameters (like ?fileDownloadPath=... or ?file=...)
            qs = urllib.parse.parse_qs(parsed_url.query)
            for key, val_list in qs.items():
                for val in val_list:
                    val_lower = val.lower()
                    for ext in valid_exts:
                        if ext in val_lower:
                            possible_file = val.split('/')[-1].split('\\')[-1]
                            if possible_file.lower().endswith(ext):
                                return clean_u, possible_file
                                
            # 3. Fallback: check if the whole URL lower case contains any extension
            for ext in valid_exts:
                if ext in clean_u.lower():
                    return clean_u, None
                    
            return None, None

        targets = []
        for url in urls:
            matched_url, filename = get_valid_document_url_and_filename(url)
            if matched_url:
                targets.append((matched_url, filename))

        if not targets:
            self._log('ok', 'No matching GeM document URLs found for download.')
            return

        self._log('info', f"Downloading {len(targets)} additional tender document(s)...")
        add_folder = os.path.join(self.filing_folder, "Additional_Tender_Documents")
        os.makedirs(add_folder, exist_ok=True)

        import urllib.request
        import ssl
        context = ssl._create_unverified_context()
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }

        for idx, (url, filename) in enumerate(targets):
            try:
                if not filename:
                    ext = ".pdf"
                    for possible_ext in ('.pdf', '.xls', '.xlsx', '.doc', '.docx', '.zip', '.csv'):
                        if possible_ext in url.lower():
                            ext = possible_ext
                            break
                    filename = f"document_{idx+1}{ext}"
                
                safe_filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
                dest_path = os.path.join(add_folder, safe_filename)

                self._log('info', f"Downloading embedded document: {safe_filename}")
                req = urllib.request.Request(url, headers=headers)
                with urllib.request.urlopen(req, timeout=20, context=context) as response:
                    with open(dest_path, 'wb') as out_file:
                        out_file.write(response.read())
                
                self._log('ok', f"Successfully downloaded: {safe_filename}")
            except Exception as e:
                self._log('warn', f"Failed to download embedded link {url}: {e}")

    def _generate_checklist(self, bid_no: str, tender_record: Dict) -> str:
        """
        Generate document checklist file.
        
        Args:
            bid_no: Tender bid number
            tender_record: Tender data dictionary
            
        Returns:
            Path to generated checklist file
        """
        checklist_path = os.path.join(self.filing_folder, 'Document_Checklist.txt')
        
        with open(checklist_path, 'w', encoding='utf-8') as f:
            f.write("=" * 60 + "\n")
            f.write("TENDER FILING DOCUMENT CHECKLIST\n")
            f.write("=" * 60 + "\n\n")
            
            f.write(f"Bid Number: {bid_no}\n")
            f.write(f"Items: {tender_record.get('items', 'N/A')}\n")
            f.write(f"Identified Category: {getattr(self, 'category', 'General')}\n")
            f.write(f"Department: {tender_record.get('dept', 'N/A')}\n")
            f.write(f"End Date: {tender_record.get('end_date', 'N/A')}\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            f.write("-" * 60 + "\n")
            f.write("DOCUMENT STATUS\n")
            f.write("-" * 60 + "\n\n")
            
            f.write(f"Total Required: {len(self.required_documents)}\n")
            f.write(f"Available: {len(self.matched_documents)}\n")
            f.write(f"Missing: {len(self.missing_documents)}\n\n")
            
            f.write("-" * 60 + "\n")
            f.write("MATCHED DOCUMENTS (✓)\n")
            f.write("-" * 60 + "\n\n")
            
            for doc_name, doc_info in self.matched_documents.items():
                req_doc = next((d for d in self.required_documents if d['name'] == doc_name), None)
                doc_category = req_doc.get('category', 'General') if req_doc else 'General'
                doc_source_file = req_doc.get('source_file', 'Tender_Document.pdf') if req_doc else 'Tender_Document.pdf'
                doc_desc = req_doc.get('description', '') if req_doc else ''
                
                if isinstance(doc_info, dict):
                    src_path = doc_info.get('path', 'N/A')
                    source = doc_info.get('source', 'N/A')
                    is_expired = doc_info.get('is_expired', False)
                    expiring_soon = doc_info.get('expiring_soon', False)
                    expiry_date = doc_info.get('expiry_date')
                    days_rem = doc_info.get('days_remaining')
                else:
                    src_path = doc_info
                    source = 'N/A'
                    is_expired = False
                    expiring_soon = False
                    expiry_date = None
                    days_rem = None
                
                if is_expired:
                    f.write(f"⚠️ {doc_name} [EXPIRED - REPLACEMENT REQUIRED!]\n")
                    f.write(f"  Expiry Date: {expiry_date} (EXPIRED)\n")
                elif expiring_soon:
                    f.write(f"⚡ {doc_name} [EXPIRING SOON]\n")
                    f.write(f"  Expiry Date: {expiry_date} ({days_rem} days remaining)\n")
                else:
                    f.write(f"✓ {doc_name}\n")
                    if expiry_date:
                        f.write(f"  Expiry Date: {expiry_date} (Valid)\n")

                f.write(f"  Source Path: {src_path}\n")
                f.write(f"  Firm Association: {source}\n")
                f.write(f"  Category: {doc_category}\n")
                f.write(f"  Requested In File: {doc_source_file}\n")
                f.write(f"  Description: {doc_desc}\n\n")
            
            f.write("-" * 60 + "\n")
            f.write("MISSING DOCUMENTS (✗)\n")
            f.write("-" * 60 + "\n\n")
            
            for doc in self.missing_documents:
                f.write(f"✗ {doc['name']}\n")
                f.write(f"  Category: {doc['category']}\n")
                f.write(f"  Requested In File: {doc.get('source_file', 'Tender_Document.pdf')}\n")
                f.write(f"  Description: {doc['description']}\n")
                f.write(f"  Required: {'Yes' if doc['required'] else 'No'}\n\n")
            
            f.write("-" * 60 + "\n")
            f.write("EMD / SECURITY DEPOSIT DETAILS\n")
            f.write("-" * 60 + "\n\n")
            
            emd = getattr(self, 'emd_details', {})
            f.write(f"EMD Required: {'Yes' if emd.get('emd_required') else 'No'}\n")
            f.write(f"EMD Amount: {emd.get('emd_amount', 'N/A')}\n")
            f.write(f"EMD Exemption Allowed: {'Yes' if emd.get('emd_exemption_allowed') else 'No'}\n")
            f.write(f"PBG/Performance Security Required: {'Yes' if emd.get('pbg_required') else 'No'}\n")
            f.write(f"PBG Percent/Amount: {emd.get('pbg_percent', 'N/A')}\n")
            f.write(f"Source Document(s): {emd.get('source_file', 'N/A')}\n")
            f.write(f"Instructions: {emd.get('details_summary', 'N/A')}\n\n")

            f.write("=" * 60 + "\n")
            f.write("END OF CHECKLIST\n")
            f.write("=" * 60 + "\n")
        
        self._log('ok', f'Generated checklist: {checklist_path}')
        return checklist_path
    
    def _generate_summary(self, bid_no: str, tender_record: Dict) -> str:
        """
        Generate filing summary file.
        
        Args:
            bid_no: Tender bid number
            tender_record: Tender data dictionary
            
        Returns:
            Path to generated summary file
        """
        summary_path = os.path.join(self.filing_folder, 'Filing_Summary.txt')
        
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write("=" * 60 + "\n")
            f.write("TENDER FILING SUMMARY\n")
            f.write("=" * 60 + "\n\n")
            
            f.write("TENDER DETAILS\n")
            f.write("-" * 60 + "\n")
            f.write(f"Bid Number: {bid_no}\n")
            f.write(f"Bid URL: {tender_record.get('bid_url', 'N/A')}\n")
            f.write(f"Ministry: {tender_record.get('ministry', 'N/A')}\n")
            f.write(f"Department: {tender_record.get('dept', 'N/A')}\n")
            f.write(f"Organisation: {tender_record.get('organisation', 'N/A')}\n")
            f.write(f"Category: {getattr(self, 'category', 'General')}\n")
            f.write(f"Items: {tender_record.get('items', 'N/A')}\n")
            f.write(f"Quantity: {tender_record.get('quantity', 'N/A')}\n")
            f.write(f"Location: {tender_record.get('location', 'N/A')}\n")
            f.write(f"Estimated Value: {tender_record.get('est_value', 'N/A')}\n")
            f.write(f"End Date: {tender_record.get('end_date', 'N/A')}\n")
            f.write(f"Bid Opening: {tender_record.get('bid_opening', 'N/A')}\n\n")
            
            f.write("EMD / SECURITY DEPOSIT DETAILS\n")
            f.write("-" * 60 + "\n")
            emd = getattr(self, 'emd_details', {})
            f.write(f"EMD Required: {'Yes' if emd.get('emd_required') else 'No'}\n")
            f.write(f"EMD Amount: {emd.get('emd_amount', 'N/A')}\n")
            f.write(f"EMD Exemption Allowed: {'Yes' if emd.get('emd_exemption_allowed') else 'No'}\n")
            f.write(f"PBG/Performance Security Required: {'Yes' if emd.get('pbg_required') else 'No'}\n")
            f.write(f"PBG Percent/Amount: {emd.get('pbg_percent', 'N/A')}\n")
            f.write(f"Source Document(s): {emd.get('source_file', 'N/A')}\n")
            f.write(f"Instructions: {emd.get('details_summary', 'N/A')}\n\n")

            f.write("FILING STATUS\n")
            f.write("-" * 60 + "\n")
            f.write(f"Filing Folder: {self.filing_folder}\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Document Completion: {len(self.matched_documents)}/{len(self.required_documents)}\n\n")
            
            f.write("NEXT STEPS\n")
            f.write("-" * 60 + "\n")
            if self.missing_documents:
                f.write("1. Arrange missing documents listed in Document_Checklist.txt\n")
                f.write("2. Review all matched documents for accuracy\n")
                f.write("3. Prepare bid submission as per tender requirements\n")
            else:
                f.write("1. Review all documents for accuracy and completeness\n")
                f.write("2. Prepare bid submission as per tender requirements\n")
            f.write("3. Submit bid before the end date\n\n")
            
            f.write("=" * 60 + "\n")
            f.write("Generated by TenderTracker Filing Workflow\n")
            f.write("=" * 60 + "\n")
        
        self._log('ok', f'Generated summary: {summary_path}')
        return summary_path

    def _extract_gem_requirements(self, pdf_text: str) -> List[str]:
        """
        Extract required documents specified in the 'Document required from seller' section of GeM bids.
        """
        import re
        if not pdf_text:
            return []
            
        # Try regex first
        pattern = r'Document\s*required\s*from\s*seller\s*\n\s*([\s\S]+?)(?=\n\s*(?:\*|Bid\s*Number|बड|सं>या|Dated|अितTर^|Additional|Consignee|\Z))'
        match = re.search(pattern, pdf_text, re.IGNORECASE)
        if match:
            raw_list = match.group(1).strip()
            # Normalize whitespace/newlines
            normalized = " ".join(raw_list.split())
            # Split by comma
            docs = [d.strip() for d in normalized.split(",") if d.strip()]
            if docs:
                self._log('ok', f'Extracted {len(docs)} GeM requirements via regex: {docs}')
                return docs

        # Fallback to LLM if enabled
        settings = db.load_settings()
        provider = settings.get('llm_provider', 'Disabled')
        if provider not in ('', 'Disabled', None):
            try:
                self._log('info', 'Attempting LLM extraction of GeM portal requirements...')
                truncated_text = pdf_text[:12000] if len(pdf_text) > 12000 else pdf_text
                prompt = f"""Extract the list of required documents specified under the section 'Document required from seller' or 'विक्रेता से मांगे गए दस्तावेज़' in the following GeM tender text.
Return ONLY a comma-separated list of the document names (e.g. "Experience Criteria, Bidder Turnover, Certificate (Requested in ATC)").
If not found, return an empty string.

Text:
{truncated_text}"""
                response_text = llm.call_llm(prompt, provider, 
                                            settings.get('llm_api_key', ''),
                                            settings.get('llm_base_url', ''),
                                            settings.get('llm_model', ''),
                                            response_json=False)
                if response_text and response_text.strip():
                    docs = [d.strip() for d in response_text.split(",") if d.strip()]
                    # Filter out non-document text if any
                    docs = [d for d in docs if len(d) < 80 and not d.startswith("Here")]
                    if docs:
                        self._log('ok', f'Extracted {len(docs)} GeM requirements via LLM: {docs}')
                        return docs
            except Exception as e:
                self._log('warn', f'LLM GeM requirements extraction failed: {e}')

        return []

    def _resolve_additional_doc_definitions(self, required_docs: List[Dict], files_text: Dict[str, str], pdf_text: str) -> List[Dict]:
        """
        Resolve generic placeholders like 'Additional Doc 1 (Requested in ATC)', 'Additional Doc 2 (Requested in ATC)',
        'Certificate (Requested in ATC)' into specific document titles by parsing downloaded additional files (ATC, BOQ, Corrigendum)
        and main PDF text.
        """
        if not required_docs:
            return required_docs

        # Identify generic placeholders that require resolution
        generic_docs = []
        for doc in required_docs:
            doc_name = doc['name']
            doc_lower = doc_name.lower()
            if any(term in doc_lower for term in ['additional doc', 'certificate (requested in atc)', 'doc 1', 'doc 2', 'doc 3', 'doc 4', 'requested in atc']):
                generic_docs.append(doc)

        if not generic_docs:
            return required_docs

        # Combine text from downloaded additional documents & main PDF
        combined_additional_text = "\n\n".join([f"=== File: {fname} ===\n{txt}" for fname, txt in files_text.items()])
        full_context_text = combined_additional_text + "\n\n" + (pdf_text[:8000] if pdf_text else "")

        if not full_context_text.strip():
            return required_docs

        resolved_map = {}

        # 1. Regex pattern matching for explicit document definitions
        # e.g., "Additional Doc 1: Land Border Sharing Declaration" or "Doc 1 (Requested in ATC) - Non Blacklisting Undertaking"
        patterns = [
            r'(?:Additional\s*)?(?:Doc|Document)\s*(\d+)[\s\(\)RequestedinATC]*[:\-–=]+\s*([^\n.,;]{3,80})',
            r'(?:Doc|Document)\s*(\d+)[\s\(\)RequestedinATC]*\s*[:\-–=]+\s*([^\n.,;]{3,80})',
            r'Certificate\s*\([^)]*ATC[^)]*\)[:\-–=]+\s*([^\n.,;]{3,80})'
        ]

        for pat in patterns:
            for match in re.finditer(pat, full_context_text, re.IGNORECASE):
                groups = match.groups()
                if len(groups) == 2:
                    doc_num, doc_def = groups
                    key = f"additional doc {doc_num}"
                else:
                    doc_def = groups[0]
                    key = "certificate (requested in atc)"
                
                doc_def_clean = re.sub(r'\s+', ' ', doc_def.strip())
                if doc_def_clean and len(doc_def_clean) > 3 and not doc_def_clean.lower().startswith('requested'):
                    if key not in resolved_map:
                        resolved_map[key] = doc_def_clean

        # 2. LLM Resolution if provider is enabled
        settings = db.load_settings()
        provider = settings.get('llm_provider', 'Disabled')
        if provider not in ('', 'Disabled', None) and len(generic_docs) > 0:
            try:
                gen_names = [d['name'] for d in generic_docs]
                prompt = f"""You are a GeM Tender Specialist. Below is text from downloaded Additional ATC Tender Documents.
The seller is asked to upload the following generic document placeholders:
{json.dumps(gen_names)}

Examine the ATC text below and extract the ACTUAL specific document title required for each placeholder.
Examples:
- "Additional Doc 1 (Requested in ATC)" -> "Land Border Sharing Declaration"
- "Additional Doc 2 (Requested in ATC)" -> "Non-Blacklisting Affidavit"
- "Additional Doc 3 (Requested in ATC)" -> "NABL Test Report for Cable"
- "Certificate (Requested in ATC)" -> "ISO 9001 Quality Certificate"

Text Excerpt:
{full_context_text[:12000]}

Return ONLY a JSON object mapping each placeholder name to its specific document title. If not found in text, omit that key or use the generic name.
JSON format:
{{
  "Additional Doc 1 (Requested in ATC)": "Land Border Sharing Declaration",
  "Additional Doc 2 (Requested in ATC)": "Non-Blacklisting Affidavit"
}}"""
                if provider == 'Local LLM (LM Studio / Ollama)':
                    self._local_llm_used = True
                res_str = llm.call_llm(prompt, provider, settings.get('llm_api_key', ''), settings.get('llm_base_url', ''), settings.get('llm_model', ''), response_json=True)
                if "</think>" in res_str:
                    res_str = res_str.split("</think>")[-1].strip()
                cleaned_json = llm.clean_json_response(res_str)
                llm_map = json.loads(cleaned_json)
                if isinstance(llm_map, dict):
                    for orig_k, resolved_v in llm_map.items():
                        if resolved_v and isinstance(resolved_v, str) and resolved_v.strip() and len(resolved_v) < 100:
                            for doc in generic_docs:
                                if orig_k.lower() in doc['name'].lower() or doc['name'].lower() in orig_k.lower():
                                    resolved_map[doc['name'].lower()] = resolved_v.strip()
            except Exception as e:
                self._log('warn', f"LLM additional doc resolution failed: {e}")

        # Apply resolved titles to required_docs
        resolved_docs = []
        for doc in required_docs:
            doc_copy = dict(doc)
            name_lower = doc_copy['name'].lower()
            
            resolved_title = None
            if name_lower in resolved_map:
                resolved_title = resolved_map[name_lower]
            else:
                for k, v in resolved_map.items():
                    if k in name_lower or name_lower in k:
                        resolved_title = v
                        break

            if resolved_title and resolved_title.lower() != doc_copy['name'].lower():
                old_name = doc_copy['name']
                doc_copy['resolved_title'] = resolved_title
                doc_copy['name'] = f"{old_name}: {resolved_title}"
                doc_copy['description'] = f"{resolved_title} (Resolved from ATC document)"
                self._log('ok', f"[ATC Resolution] Resolved '{old_name}' ➔ '{doc_copy['name']}'")

            resolved_docs.append(doc_copy)

        return resolved_docs

    def _generate_auto_declarations(self, active_firm: str, tender_record: Dict) -> List[str]:
        """
        Auto-generate compliance declarations (MII, Land Border, Non-Blacklisting, Bidder Undertaking, Bid Securing).
        Saves generated files into '02_Generated_Declarations' subfolder.
        """
        if not self.filing_folder or not os.path.exists(self.filing_folder):
            return []

        decl_folder = os.path.join(self.filing_folder, "02_Generated_Declarations")
        try:
            os.makedirs(decl_folder, exist_ok=True)
        except Exception as e:
            self._log('warn', f"Failed to create 02_Generated_Declarations folder: {e}")
            return []

        settings = db.load_settings()
        firms = settings.get('firms', [])
        firm_info = next((f for f in firms if f.get('name') == active_firm), {})
        if not firm_info and firms:
            firm_info = firms[0]

        context = {
            "bid_no": tender_record.get('bid_no', 'N/A'),
            "category": getattr(self, 'category', 'General'),
            "firm_name": active_firm or "Unnamed Firm",
            "firm_address": firm_info.get('address', 'N/A'),
            "department": tender_record.get('dept', 'Government of India (GeM Portal)'),
            "date": datetime.now().strftime("%d-%m-%Y"),
            "signatory_name": firm_info.get('signatory_name', 'Authorized Representative'),
            "signatory_designation": firm_info.get('signatory_designation', 'Proprietor / Authorized Signatory'),
            "local_content_percentage": "50",
            "local_content_location": firm_info.get('address', 'Works Premises')
        }

        template_types = [
            "bidder_undertaking",
            "declaration",
            "affidavit",
            "mii_certificate",
            "land_border_declaration",
            "bid_securing_declaration"
        ]

        generated_files = []
        for t_type in template_types:
            try:
                ok, path_out, warn = template_generator.generate_document(decl_folder, t_type, context)
                if ok and path_out:
                    generated_files.append(path_out)
                    self._log('ok', f"Generated compliance declaration: {os.path.basename(path_out)}")
            except Exception as e:
                self._log('warn', f"Failed to generate template declaration '{t_type}': {e}")

        return generated_files

    def _calculate_filing_readiness(self) -> Dict:
        """
        Calculate pre-filing readiness score (0-100%) and disqualification risk audit.
        """
        total_req = len(self.required_documents)
        matched_req = len(self.matched_documents)
        missing_req = len(self.missing_documents)

        # Check for expired matched documents
        expired_count = 0
        expiring_soon_count = 0
        for doc_info in self.matched_documents.values():
            if isinstance(doc_info, dict):
                if doc_info.get('is_expired'):
                    expired_count += 1
                elif doc_info.get('expiring_soon'):
                    expiring_soon_count += 1

        # Check for oversized documents
        oversized_count = 0
        for v in self.validation_results.values():
            if isinstance(v, dict) and v.get('warnings'):
                if any('exceeds GEM limit' in w for w in v['warnings']):
                    oversized_count += 1

        # Match score (60% weight)
        match_score = (matched_req / total_req * 60.0) if total_req > 0 else 60.0

        # Expiry penalty (up to -30%)
        expiry_penalty = expired_count * 15.0

        # Oversized penalty (up to -10%)
        oversized_penalty = oversized_count * 5.0

        # Final score
        raw_score = max(0.0, min(100.0, match_score + 40.0 - expiry_penalty - oversized_penalty))
        score = round(raw_score, 1)

        status_label = "READY TO FILE" if score >= 85 and expired_count == 0 else ("ACTION REQUIRED" if score < 60 or expired_count > 0 else "PARTIALLY READY")

        return {
            "score": score,
            "status_label": status_label,
            "matched_count": matched_req,
            "missing_count": missing_req,
            "expired_count": expired_count,
            "expiring_soon_count": expiring_soon_count,
            "oversized_count": oversized_count,
        }

    def _generate_excel_checklist(self, bid_no: str, tender_record: Dict) -> str:
        """Export Filing_Checklist.xlsx with formatting, status badges, and details."""
        if not self.filing_folder or not os.path.exists(self.filing_folder):
            return ""

        excel_path = os.path.join(self.filing_folder, "Filing_Checklist.xlsx")
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Filing Checklist"

            # Header info
            ws.append(["TENDER FILING CHECKLIST & READINESS REPORT"])
            ws.append(["Bid Number", bid_no, "End Date", tender_record.get('end_date', 'N/A')])
            ws.append(["Category", getattr(self, 'category', 'General'), "Department", tender_record.get('dept', 'N/A')])
            ws.append([])

            headers = ["Status", "Document Name", "Category", "Source Path", "Expiry Date", "Requested In", "Description"]
            ws.append(headers)

            # Style header row
            header_fill = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
            header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=5, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Matched documents
            valid_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            expired_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            warn_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

            row_idx = 6
            for doc_name, doc_info in self.matched_documents.items():
                req_doc = next((d for d in self.required_documents if d['name'] == doc_name), None)
                cat = req_doc.get('category', 'General') if req_doc else 'General'
                src_file = req_doc.get('source_file', 'Tender_Document.pdf') if req_doc else 'Tender_Document.pdf'
                desc = req_doc.get('description', '') if req_doc else ''

                if isinstance(doc_info, dict):
                    src_path = doc_info.get('path', 'N/A')
                    is_exp = doc_info.get('is_expired', False)
                    exp_soon = doc_info.get('expiring_soon', False)
                    exp_date = doc_info.get('expiry_date') or 'N/A'
                else:
                    src_path = str(doc_info)
                    is_exp, exp_soon, exp_date = False, False, 'N/A'

                if is_exp:
                    status = "[EXPIRED]"
                    fill = expired_fill
                elif exp_soon:
                    status = "[EXPIRING SOON]"
                    fill = warn_fill
                else:
                    status = "[VALID]"
                    fill = valid_fill

                ws.append([status, doc_name, cat, src_path, exp_date, src_file, desc])
                for col_idx in range(1, 8):
                    ws.cell(row=row_idx, column=col_idx).fill = fill
                row_idx += 1

            # Missing documents
            for doc in self.missing_documents:
                ws.append(["[MISSING]", doc['name'], doc.get('category', 'General'), "N/A", "N/A", doc.get('source_file', 'Tender_Document.pdf'), doc.get('description', '')])
                for col_idx in range(1, 8):
                    ws.cell(row=row_idx, column=col_idx).fill = expired_fill
                row_idx += 1

            wb.save(excel_path)
            self._log('ok', f"Generated Excel checklist: {excel_path}")
            return excel_path

        except Exception as e:
            self._log('warn', f"Failed to generate Filing_Checklist.xlsx: {e}")
            return ""

