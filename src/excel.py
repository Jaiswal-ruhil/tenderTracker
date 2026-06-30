import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XL_HEADERS = [
    "S.No", "Bid No", "Bid URL",
    "Ministry", "Department / Buyer", "Organisation", "Office",
    "Category", "Items", "Quantity",
    "Consignee / Location", "Contract Duration",
    "Estimated Value (Rs)", "Evaluation Method",
    "Bid Type", "Bid to RA",
    "EMD Required", "ePBG Required",
    "MII Compliance", "MSE Purchase Preference",
    "MSE Relaxation", "Startup Relaxation",
    "Min Turnover (Lakhs)", "Experience Required (Yrs)",
    "Bid Opening Date", "Start Date", "End Date",
    "Entry Date", "Filing Status", "Remarks",
]
COL_WIDTHS = [5,22,40,28,30,24,20,22,30,8,24,14,16,20,14,10,10,10,10,10,10,10,14,14,18,18,18,14,15,22]

THIN     = Side(style="thin")
BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", start_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DAT_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill("solid", start_color="DCE6F1")

def financial_year(dt):
    return f"{dt.year}-{str(dt.year+1)[-2:]}" if dt.month >= 4 \
           else f"{dt.year-1}-{str(dt.year)[-2:]}"

def xl_path(folder, fy, pattern=None):
    if not pattern:
        pattern = "GEM_Tenders_FY_{fy}"
    now_str = datetime.now().strftime("%d-%m-%Y")
    filename = pattern.replace("{fy}", fy).replace("{date}", now_str)
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"
    return os.path.join(folder, filename)

def ensure_workbook(path):
    if os.path.exists(path): return
    wb = Workbook(); ws = wb.active; ws.title = "Tenders"
    ws.append(XL_HEADERS)
    for col in range(1, len(XL_HEADERS)+1):
        c = ws.cell(row=1, column=col)
        c.font=HDR_FONT; c.fill=HDR_FILL; c.border=BORDER
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.row_dimensions[1].height=32
    for i,w in enumerate(COL_WIDTHS,1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A2"; wb.save(path)

def xl_append(path, rows):
    wb = load_workbook(path); ws = wb["Tenders"]; snos=[]
    for rec in rows:
        sno = ws.max_row
        ws.append([
            sno,
            rec.get("bid_no",""),      rec.get("bid_url",""),
            rec.get("ministry",""),    rec.get("dept",""),
            rec.get("organisation",""),rec.get("office",""),
            rec.get("category",""),    rec.get("items",""),
            rec.get("quantity",""),    rec.get("location",""),
            rec.get("contract_dur",""),rec.get("est_value",""),
            rec.get("eval_method",""), rec.get("bid_type",""),
            rec.get("bid_to_ra",""),   rec.get("emd",""),
            rec.get("epbg",""),        rec.get("mii",""),
            rec.get("mse_pref",""),    rec.get("mse_relax",""),
            rec.get("startup_relax",""),rec.get("min_turnover",""),
            rec.get("exp_years",""),   rec.get("bid_opening",""),
            rec.get("start_date",""),  rec.get("end_date",""),
            datetime.now().strftime("%d-%m-%Y %H:%M"),
            rec.get("filing_status","Not Filed"),
            rec.get("remarks",""),
        ])
        nr=ws.max_row; alt=sno%2==0
        for col in range(1,len(XL_HEADERS)+1):
            c=ws.cell(row=nr,column=col)
            c.font=DAT_FONT; c.border=BORDER
            c.alignment=Alignment(vertical="center",wrap_text=True)
            if alt: c.fill=ALT_FILL
        ws.row_dimensions[nr].height=20; snos.append(sno)
    wb.save(path); return snos
