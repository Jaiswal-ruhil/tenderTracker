import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

import unittest
from datetime import datetime
import openpyxl
import excel

TEST_XL_FILE = os.path.join(os.path.dirname(__file__), "test_tenders.xlsx")

class TestExcel(unittest.TestCase):
    def setUp(self):
        if os.path.exists(TEST_XL_FILE):
            os.remove(TEST_XL_FILE)

    def tearDown(self):
        if os.path.exists(TEST_XL_FILE):
            os.remove(TEST_XL_FILE)

    def test_financial_year(self):
        # Boundaries check
        # April 1st, 2026 -> 2026-27
        dt1 = datetime(2026, 4, 1)
        self.assertEqual(excel.financial_year(dt1), "2026-27")
        
        # March 31st, 2026 -> 2025-26
        dt2 = datetime(2026, 3, 31)
        self.assertEqual(excel.financial_year(dt2), "2025-26")
        
        # December 31st, 2026 -> 2026-27
        dt3 = datetime(2026, 12, 31)
        self.assertEqual(excel.financial_year(dt3), "2026-27")
        
        # January 1st, 2027 -> 2026-27
        dt4 = datetime(2027, 1, 1)
        self.assertEqual(excel.financial_year(dt4), "2026-27")

    def test_xl_path(self):
        folder = "D:/Temp"
        fy = "2026-27"
        
        # Test default pattern
        expected = os.path.join(folder, "GEM_Tenders_FY_2026-27.xlsx")
        self.assertEqual(excel.xl_path(folder, fy), expected)
        
        # Test custom pattern with {fy}
        expected_custom1 = os.path.join(folder, "Tenders_FY_2026-27.xlsx")
        self.assertEqual(excel.xl_path(folder, fy, pattern="Tenders_FY_{fy}"), expected_custom1)
        
        # Test custom pattern with {date}
        now_str = datetime.now().strftime("%d-%m-%Y")
        expected_custom2 = os.path.join(folder, f"Tenders_Export_{now_str}.xlsx")
        self.assertEqual(excel.xl_path(folder, fy, pattern="Tenders_Export_{date}"), expected_custom2)
        
        # Test pattern without .xlsx extension
        expected_custom3 = os.path.join(folder, f"Tenders_Export_{now_str}_2026-27.xlsx")
        self.assertEqual(excel.xl_path(folder, fy, pattern="Tenders_Export_{date}_{fy}"), expected_custom3)

    def test_ensure_workbook(self):
        # Should create the file
        excel.ensure_workbook(TEST_XL_FILE)
        self.assertTrue(os.path.exists(TEST_XL_FILE))
        
        # Check workbook properties
        wb = openpyxl.load_workbook(TEST_XL_FILE)
        self.assertIn("Tenders", wb.sheetnames)
        ws = wb["Tenders"]
        
        # Check headers match XL_HEADERS
        headers = [ws.cell(row=1, column=col).value for col in range(1, len(excel.XL_HEADERS) + 1)]
        self.assertEqual(headers, excel.XL_HEADERS)
        
        # Header formatting check
        c = ws.cell(row=1, column=1)
        self.assertTrue(c.font.bold)
        self.assertEqual(c.fill.start_color.rgb, "001F4E79") # HDR_FILL starts with '1F4E79' (AARRGGBB in openpyxl)

    def test_xl_append(self):
        excel.ensure_workbook(TEST_XL_FILE)
        
        recs = [
            {
                "bid_no": "GEM/2026/B/100",
                "bid_url": "https://example.com/100",
                "ministry": "Mines",
                "items": "Laptop",
                "quantity": "5",
                "location": "Delhi"
            },
            {
                "bid_no": "GEM/2026/B/200",
                "bid_url": "https://example.com/200",
                "ministry": "Coal",
                "items": "Scanner",
                "quantity": "2",
                "location": "Kolkata"
            }
        ]
        
        snos = excel.xl_append(TEST_XL_FILE, recs)
        self.assertEqual(snos, [1, 2])
        
        # Read back and verify values
        wb = openpyxl.load_workbook(TEST_XL_FILE)
        ws = wb["Tenders"]
        self.assertEqual(ws.max_row, 3) # 1 header row + 2 data rows
        
        # Check first row values
        self.assertEqual(ws.cell(row=2, column=1).value, 1) # S.No
        self.assertEqual(ws.cell(row=2, column=2).value, "GEM/2026/B/100") # Bid No
        self.assertEqual(ws.cell(row=2, column=3).value, "https://example.com/100") # Bid URL
        self.assertEqual(ws.cell(row=2, column=4).value, "Mines") # Ministry
        self.assertEqual(ws.cell(row=2, column=9).value, "Laptop") # Items (Col 9 is Items)
        
        # Check second row values
        self.assertEqual(ws.cell(row=3, column=1).value, 2)
        self.assertEqual(ws.cell(row=3, column=2).value, "GEM/2026/B/200")
        
        # Alternate fill color check:
        # Row 2 (SNo 1) is odd -> no alternate fill (or check is odd/even check in xl_append)
        # In xl_append: `alt = sno % 2 == 0`
        # Row 2 has S.No = 1. So alt = 1 % 2 == 0 -> False (No alternate fill)
        # Row 3 has S.No = 2. So alt = 2 % 2 == 0 -> True (Alternate fill applied!)
        
        c2 = ws.cell(row=2, column=1) # SNo 1
        c3 = ws.cell(row=3, column=1) # SNo 2
        
        self.assertEqual(c2.fill.fill_type, None) # No fill
        self.assertEqual(c3.fill.start_color.rgb, "00DCE6F1") # ALT_FILL is 'DCE6F1'

if __name__ == '__main__':
    unittest.main()
